#!/usr/bin/env python3
"""朝钉钉群推送续费周报：简报 + 指定小组跟进清单（含 LP 执行问题）。

用法:
  python dingtalk_push.py --report <报告.md> --intent-excel <意向.xlsx> --quality-excel <质量.xlsx> [--test]
"""
import argparse
import json
import sys
import time
import urllib.request
from pathlib import Path
from collections import defaultdict, Counter
from typing import Any

import openpyxl

# 多群推送：从 config 读取所有 webhook keys
import config
DEFAULT_KEYS = config.get("dingtalk_webhook_keys")


class DingTalkPusher:
    def __init__(self, webhook_key: str):
        self.webhook_key = webhook_key

    def send_markdown(self, title: str, text: str) -> dict[str, Any]:
        url = f"https://oapi.dingtalk.com/robot/send?access_token={self.webhook_key}"
        data = json.dumps({
            "msgtype": "markdown",
            "markdown": {"title": title, "text": text}
        }, ensure_ascii=False)
        req = urllib.request.Request(
            url, data=data.encode("utf-8"),
            headers={"Content-Type": "application/json; charset=utf-8"},
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read().decode("utf-8")
        return json.loads(body)

    def send_text(self, content: str) -> dict[str, Any]:
        url = f"https://oapi.dingtalk.com/robot/send?access_token={self.webhook_key}"
        data = json.dumps({"msgtype": "text", "text": {"content": content}}, ensure_ascii=False)
        req = urllib.request.Request(url, data=data.encode("utf-8"),
                                     headers={"Content-Type": "application/json; charset=utf-8"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))


# 小组配置
TARGET_GROUPS = {
    "港澳1组": "叶素姜",
    "港澳2组": "向梦清",
    "港澳组": "谢昌馥",
}

# LP 问题三分类
PROBLEM_CATEGORIES = {
    "沟通执行": [
        "模板轰炸(>=5条)", "模板偏多(3-4条)", "打卡任务干扰续费对话",
        "家长零回复(单向广播)", "家长几乎无互动",
        "家长拒绝后未做有效挽救", "缺乏个性化沟通",
    ],
    "能力": [
        "未使用退费保障打消顾虑", "未做价格对比/锚点",
        "使用情绪施压话术(考核/指标)", "未做学习痛点/进步点挖掘",
        "转介绍推荐喧宾夺主",
    ],
    # "无对话内容" 不再算 LP 执行问题（原始数据缺失，非 LP 责任）
    "_data_issues": ["无对话内容"],
}

# 严重程度标记
SEVERITY = {
    "无对话内容": "🔴", "家长零回复(单向广播)": "🔴",
    "家长拒绝后未做有效挽救": "🔴", "使用情绪施压话术(考核/指标)": "🔴",
    "模板轰炸(>=5条)": "🟡", "打卡任务干扰续费对话": "🟡",
    "家长几乎无互动": "🟡", "未使用退费保障打消顾虑": "🟡",
    "未做价格对比/锚点": "🟡", "未做学习痛点/进步点挖掘": "🟡",
    "模板偏多(3-4条)": "🟢", "转介绍推荐喧宾夺主": "🟢",
    "缺乏个性化沟通": "🟡",
}


def read_intent_excel(path: str) -> dict:
    """读取续费意向分析 Excel"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["续费意向盘"]
    # Read headers from row 1
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    header_map = {}
    for i, h in enumerate(headers):
        header_map[h] = i

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [c for c in row]
        if not vals or not vals[header_map.get("学员ID", 1) if "学员ID" in header_map else 1]:
            continue
        rows.append({
            "student_id": str(vals[header_map.get("学员ID", 1)]).strip().rstrip(".0") if header_map.get("学员ID", 1) < len(vals) else "",
            "lp": str(vals[header_map.get("归属LP", 2)]).strip() if header_map.get("归属LP", 2) < len(vals) else "",
            "group": str(vals[header_map.get("小组", 3)]).strip() if header_map.get("小组", 3) < len(vals) else "",
            "level": str(vals[header_map.get("意向等级", 5)]).strip() if header_map.get("意向等级", 5) < len(vals) else "",
            "score": int(float(vals[header_map.get("综合得分", 6)])) if header_map.get("综合得分", 6) < len(vals) else 0,
            "pos_signals": str(vals[header_map.get("正向信号", 10)]).strip() if header_map.get("正向信号", 10) < len(vals) else "",
            "risk_signals": str(vals[header_map.get("风险信号", 11)]).strip() if header_map.get("风险信号", 11) < len(vals) else "",
            "action": str(vals[header_map.get("跟进行动建议", 13)]).strip() if header_map.get("跟进行动建议", 13) < len(vals) else "",
        })
    wb.close()
    return {"rows": rows}


def read_quality_excel(path: str) -> dict:
    """读取销售质量分析 Excel，返回 LP → 问题列表 映射"""
    wb = openpyxl.load_workbook(path, data_only=True)
    lp_problems = defaultdict(lambda: defaultdict(int))  # lp -> {problem: count}
    lp_total = Counter()

    # Sheet 1: 学员综合分析 (per-student problem data)
    ws1 = wb["学员综合分析"]
    headers1 = [str(c.value).strip() if c.value else "" for c in ws1[1]]
    # Try multiple possible header names
    lp_col = None
    for name in ["归属LP", "LP"]:
        try: lp_col = headers1.index(name); break
        except ValueError: pass
    problems_col = None
    for name in ["检测问题", "LP问题点"]:
        try: problems_col = headers1.index(name); break
        except ValueError: pass

    if lp_col is not None and problems_col is not None:
        for row in ws1.iter_rows(min_row=2, values_only=True):
            lp_name = str(row[lp_col]).strip() if row[lp_col] else ""
            prob_text = str(row[problems_col]).strip() if row[problems_col] else ""
            if not lp_name or not prob_text or prob_text in ("None", "[]", ""):
                continue
            lp_total[lp_name] += 1
            # Parse problem list: "[问题1, 问题2, ...]"
            for prob in prob_text.strip("[]").split(","):
                prob = prob.strip().strip("'\"")
                if prob:
                    lp_problems[lp_name][prob] += 1

    wb.close()
    return {"lp_problems": dict(lp_problems), "lp_total": dict(lp_total)}


def build_brief(date_start: str, date_end: str, group_data: list[dict], feishu_url: str = "", pool_data: list = None) -> str:
    """构建简报 Markdown — 优化展示"""
    lines = []
    # ── 标题区 ──
    lines.append(f"# 海外续费意向周报")
    lines.append(f"### {date_start} — {date_end}")
    lines.append("")

    total = sum(g["学员数"] for g in group_data)
    high_total = sum(g["高意向"] for g in group_data)
    risk_total = sum(g["风险"] for g in group_data)
    med_total = sum(g.get("中意向", 0) for g in group_data)
    low_total = sum(g["观望"] for g in group_data)

    # ── 意向总览卡片 ──
    h_pct = high_total / total * 100
    r_pct = risk_total / total * 100
    # 整体健康度判断
    if h_pct >= 15:
        health = "🟢 健康"
    elif h_pct >= 8:
        health = "🟡 一般"
    else:
        health = "🔴 待提升"

    lines.append(f"## 意向总览  `{health}`")
    lines.append("")
    lines.append(f"> 🔥 **高意向 {high_total}人** ({h_pct:.1f}%)　🟡 中意向 {med_total}人　🔵 观望 {low_total}人　🔴 风险 {risk_total}人 ({r_pct:.1f}%)")
    lines.append("")

    # ── 小组排名（港澳三组高亮）──
    target_names = set(TARGET_GROUPS.keys())
    all_groups = sorted(group_data, key=lambda x: (x["小组"] not in target_names, -x["学员数"]))

    lines.append(f"## 小组排名")
    lines.append("")
    lines.append(f"| 小组 | 学员 | 🔥高 | 高意向率 | 🔴险 | 风险率 | 组长 |")
    lines.append(f"|------|:---:|:--:|:------:|:--:|:-----:|------|")
    for g in all_groups:
        in_target = g["小组"] in target_names
        prefix = "**" if in_target else ""
        suffix = "**" if in_target else ""
        tl = TARGET_GROUPS.get(g["小组"], "—")
        h_rate = f"{g['高意向']/g['学员数']*100:.1f}%" if g["学员数"] else "0%"
        r_rate = f"{g['风险']/g['学员数']*100:.1f}%" if g["学员数"] else "0%"
        # 风险率着色
        risk_pct = g["风险"] / max(g["学员数"], 1) * 100
        if risk_pct >= 10:
            r_display = f"<font color=#FF4D4F>{r_rate}</font>"
        elif risk_pct >= 5:
            r_display = f"<font color=#FAAD14>{r_rate}</font>"
        else:
            r_display = r_rate
        lines.append(f"| {prefix}{g['小组']}{suffix} | {g['学员数']} | {g['高意向']} | {h_rate} | {g['风险']} | {r_display} | {tl} |")
    lines.append("")

    # ── 关注信号 ──
    lines.append("---")
    lines.append(f"### ⚡ 本周行动")
    lines.append("")
    max_risk_group = max(all_groups, key=lambda x: x["风险"] / max(x["学员数"], 1))
    lines.append(f"- ⚠️ **{max_risk_group['小组']}** 风险率最高 ({max_risk_group['风险']/max(max_risk_group['学员数'],1)*100:.1f}%)")
    lines.append(f"- 📌 高意向学员 **48h内电话成交**，不要发模板消息")
    lines.append(f"- 📌 风险学员 **主管逐一审核挽留**")
    lines.append(f"- 📋 下方为续费过程分析表，明细见飞书链接")
    
    # ── 飞书链接（必须在过程分析表前，避免被截断）──
    if feishu_url:
        lines.append(f"- 📊 [意向客户明细 + 过程分析表（飞书表格）]({feishu_url})")
    
    # ── 指标说明 ──
    if pool_data:
        lines.append("")
        lines.append("**指标说明**  ")
        lines.append("- **>7天**：核心池中最近沟通超7天学员占比（分母=核心池）。越高跟进节奏越差。  ")
        lines.append("- **断联率**：曾覆盖但超7天未联系占比（分母=有执行动作核心池）。跟过但断了的比例。  ")
        lines.append("")
    
    # ── 续费过程分析表 v4: 核心池展开 + 非核心占比 ──
    if pool_data:
        lines.append("")
        lines.append("## 续费过程分析")
        lines.append("")
        lines.append("| 团队 | 池子 | 学员 | 签单率 | 未续 | 覆盖 | 高意向占比 | 中意向占比 | 低意向占比 | 风险占比 | >7天 | 断联率 | 退费占比 | 结课占比 |")
        lines.append("|------|------|:---:|:-----:|:---:|:---:|:--------:|:--------:|:--------:|:------:|:---:|:---:|:------:|:------:|")
        for r in pool_data:
            team = r.get("团队", "")
            pool = r.get("池子", "")
            core = r.get("学员", 0)
            sign = r.get("签单率", "-")
            uns = r.get("未续", 0)
            cov = r.get("覆盖率", "-")
            hi = r.get("高", "-"); mi = r.get("中", "-")
            lo = r.get("低", "-"); ri_val = r.get("风险", "-")
            gt7 = r.get(">7天", "-")
            dc = r.get("断联率", "-")
            rf = r.get("退费", "0%"); fh = r.get("结课", "0%")

            team_display = f"**{team}**" if team else ""
            gt7_disp = gt7
            try:
                if float(gt7.replace("%","")) >= 30: gt7_disp = f"<font color=#FF4D4F>**{gt7}**</font>"
                elif float(gt7.replace("%","")) >= 20: gt7_disp = f"<font color=#FAAD14>{gt7}</font>"
            except: pass
            cov_disp = cov
            try:
                if float(cov.replace("%","")) < 25: cov_disp = f"<font color=#FF4D4F>{cov}</font>"
                elif float(cov.replace("%","")) < 40: cov_disp = f"<font color=#FAAD14>{cov}</font>"
            except: pass

            dc_disp = dc
            try:
                if float(dc.replace("%","")) >= 50: dc_disp = f"<font color=#FF4D4F>**{dc}**</font>"
                elif float(dc.replace("%","")) >= 30: dc_disp = f"<font color=#FAAD14>{dc}</font>"
            except: pass

            lines.append(f"| {team_display} | {pool} | {core} | {sign} | {uns} | {cov_disp} | {hi} | {mi} | {lo} | {ri_val} | {gt7_disp} | {dc_disp} | {rf} | {fh} |")
        lines.append("")
    
    lines.append("")

    return "\n".join(lines)


def build_group_detail(
    group_name: str, leader: str,
    intent_rows: list[dict],
    lp_problems: dict, lp_total: dict,
    quality_rows: list[dict] = None,
) -> str:
    """构建单个组的跟进清单 — 优化展示"""
    lines = []

    # ── 组头 ──
    group_students = [r for r in intent_rows if r["group"] == group_name]
    total = len(group_students)
    high = [r for r in group_students if "HIGH" in r["level"] or "高" in r["level"]]
    med  = [r for r in group_students if "MED" in r["level"] or "中" in r["level"]]
    risk = [r for r in group_students if "RISK" in r["level"] or "风险" in r["level"]]

    lines.append(f"# 📋 {group_name}")
    lines.append(f"### 组长：**{leader}**　　👥{total}人　　{len(high)}高 / {len(med)}中 / {len(risk)}险")
    lines.append("")

    # ── 高意向 ──
    high_sorted = sorted(high, key=lambda x: -x["score"])[:15]
    if high_sorted:
        lines.append(f"### 🔥 高意向 · 48h跟进 ({len(high)}人)")
        lines.append("")
        lines.append(f"| 学员ID | LP | 分 | 关键信号 |")
        lines.append(f"|--------|-----|:--:|---------|")
        for r in high_sorted:
            sid = r["student_id"].rstrip(".0")
            lp = r["lp"]
            sc = r["score"]
            sig = r["pos_signals"][:36] if r["pos_signals"] else "—"
            lines.append(f"| {sid} | {lp} | {sc} | {sig} |")
        lines.append("")

    # ── 风险学员 ──
    risk_sorted = sorted(risk, key=lambda x: x["score"])
    if risk_sorted:
        lines.append(f"### 🔴 风险学员 · 主管挽留 ({len(risk)}人)")
        lines.append("")
        lines.append(f"| 学员ID | LP | 风险信号 |")
        lines.append(f"|--------|-----|---------|")
        for r in risk_sorted:
            sid = r["student_id"].rstrip(".0")
            lp = r["lp"]
            sig = r["risk_signals"][:40] if r["risk_signals"] else "—"
            lines.append(f"| {sid} | {lp} | {sig} |")
        lines.append("")

    # ── LP 执行问题 ──
    group_lps = set(r["lp"] for r in group_students if r["lp"])
    lp_issues = []
    for lp_name in group_lps:
        probs = lp_problems.get(lp_name, {})
        risk_count = sum(1 for r in risk if r["lp"] == lp_name)
        exec_problems = {p: c for p, c in probs.items()
                         if p in PROBLEM_CATEGORIES["沟通执行"]}
        capa_problems = {p: c for p, c in probs.items()
                         if p in PROBLEM_CATEGORIES["能力"]}
        total_problems = len(exec_problems) + len(capa_problems)
        if total_problems > 0 or risk_count > 0:
            lp_issues.append({
                "lp": lp_name, "risk_count": risk_count,
                "exec": exec_problems, "capa": capa_problems, "total": total_problems,
            })

    lp_issues.sort(key=lambda x: (-x["risk_count"], -x["total"]))

    if lp_issues:
        lines.append(f"### 📉 LP 执行诊断")
        lines.append("")
        lines.append(f"| LP | 风险 | 沟通执行问题 | 能力问题 |")
        lines.append(f"|-----|:---:|------------|---------|")
        for li in lp_issues[:15]:
            risk_tag = f"<font color=#FF4D4F>**{li['risk_count']}**</font>" if li['risk_count'] > 0 else "0"
            exec_str = "、".join(f"{SEVERITY.get(p,'')}{p}({c})" for p, c in li["exec"].items()) or "—"
            capa_str = "、".join(f"{SEVERITY.get(p,'')}{p}({c})" for p, c in li["capa"].items()) or "—"
            if len(exec_str) > 60:
                exec_str = exec_str[:57] + "..."
            if len(capa_str) > 60:
                capa_str = capa_str[:57] + "..."
            lines.append(f"| {li['lp']} | {risk_tag} | {exec_str} | {capa_str} |")
        lines.append("")

    lines.append("---")
    lines.append(f"🔴严重　🟡中等　🟢轻度　　沟通执行=流程纪律　能力=话术技能")
    lines.append("")

    return "\n".join(lines)


def get_group_summary_from_report(group_data: list[dict]) -> list[dict]:
    """从 group_stats 报告中提取小组汇总数据（直接在这里计算）"""
    return group_data


def build_sid_group_map(raw_path: str) -> dict:
    """从原始 BI 导出 Excel 建立 学员ID→小组 映射"""
    import pandas as pd
    df = pd.read_excel(raw_path, dtype=str)
    header_idx = None
    for i in range(min(15, len(df))):
        row_vals = [str(v).strip() for v in df.iloc[i].values]
        if "学员ID" in row_vals and sum(1 for v in row_vals if "小组" in str(v) or "LP" in str(v)) >= 1:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 4
    df.columns = [str(c).strip() for c in df.iloc[header_idx].values]
    df = df.iloc[header_idx + 1:].reset_index(drop=True)
    sid_col = next((c for c in df.columns if "学员ID" in str(c)), "学员ID")
    group_col = next((c for c in df.columns if "小组" in str(c)), "小组")
    result = {}
    for _, row in df.iterrows():
        sid = str(row[sid_col]).strip()
        gp = str(row[group_col]).strip()
        if sid and sid != "nan" and gp and gp != "nan":
            result[sid] = gp
    return result


def main():
    parser = argparse.ArgumentParser(description="钉钉推送续费周报（简报+分组明细+LP问题）")
    parser.add_argument("--report", help="group_report Markdown 路径（用于提取小组统计）")
    parser.add_argument("--intent-excel", required=True, help="续费意向分析 Excel 路径")
    parser.add_argument("--quality-excel", required=True, help="销售质量分析 Excel 路径")
    parser.add_argument("--raw-excel", help="原始 BI 导出 Excel 路径（用于补充小组列映射）")
    parser.add_argument("--date-start", help="报告周期起始 (如 2026-05-12)")
    parser.add_argument("--date-end", help="报告周期结束 (如 2026-05-20)")
    parser.add_argument("--feishu-url", default="", help="飞书表格链接（嵌入简报底部）")
    parser.add_argument("--pool-excel", default="", help="池子统计 Excel 路径")
    parser.add_argument("--key", nargs="*", default=DEFAULT_KEYS,
                        help="钉钉 Webhook Key（可多个），默认从 config 读取全部")
    parser.add_argument("--test", action="store_true", help="仅打印内容，不发送")
    parser.add_argument("--msg", type=int, default=0, help="仅发送指定消息编号(1-4)，0=全部")
    args = parser.parse_args()

    # Read data
    intent_data = read_intent_excel(args.intent_excel)

    # Build SID → group map from raw data if provided
    if args.raw_excel:
        sid_group = build_sid_group_map(args.raw_excel)
        for r in intent_data["rows"]:
            sid = r["student_id"]
            # Try with and without .0 suffix
            if sid in sid_group:
                r["group"] = sid_group[sid]
            elif sid + ".0" in sid_group:
                r["group"] = sid_group[sid + ".0"]
            elif sid.rstrip(".0") in sid_group:
                r["group"] = sid_group[sid.rstrip(".0")]

    quality_data = read_quality_excel(args.quality_excel)

    print(f"读取意向数据: {len(intent_data['rows'])} 学员")
    print(f"读取质量数据: {len(quality_data['lp_problems'])} 位 LP 有问题记录")

    # Build group summary
    group_stats = defaultdict(lambda: {"小组": "", "学员数": 0, "高意向": 0, "中意向": 0, "观望": 0, "风险": 0})
    for r in intent_data["rows"]:
        g = r["group"] or "未知组"
        group_stats[g]["小组"] = g
        group_stats[g]["学员数"] += 1
        lv = r["level"]
        if "HIGH" in lv or "高" in lv:
            group_stats[g]["高意向"] += 1
        elif "MED" in lv or "中" in lv:
            group_stats[g]["中意向"] += 1
        elif "RISK" in lv or "风险" in lv:
            group_stats[g]["风险"] += 1
        else:
            group_stats[g]["观望"] += 1

    group_list = list(group_stats.values())

    date_start = args.date_start or "上周一"
    date_end = args.date_end or "昨天"

    messages = []

    # Read pool data (process report v3: 15-col format matching process_report.py output)
    pool_data = None
    if args.pool_excel and Path(args.pool_excel).exists():
        try:
            pw = openpyxl.load_workbook(args.pool_excel, data_only=True)
            pws = pw.active
            pool_data = []
            for row in pws.iter_rows(min_row=2, values_only=True):
                if row[0] or row[1]:
                    pool_data.append({
                        "团队": str(row[0] or "").strip(),
                        "池子": str(row[1] or "").strip(),
                        "学员": row[2] or 0,
                        "签单率": str(row[3] or ""),
                        "未续": row[4] or 0,
                        "覆盖率": str(row[5] or ""),
                        "高": str(row[6] or ""), "中": str(row[7] or ""),
                        "低": str(row[8] or ""), "风险": str(row[9] or ""),
                        ">7天": str(row[10] or ""),
                        "断联率": str(row[11] or ""),
                        "退费": str(row[12] or ""), "结课": str(row[13] or ""),
                    })
            pw.close()
        except Exception as e:
            print(f"读取过程报告失败: {e}")

    # Message 1: Brief
    brief = build_brief(date_start, date_end, group_list, args.feishu_url, pool_data)
    messages.append(("海外续费意向周报", brief))

    # Messages 2-4: Group details
    for group_name, leader in TARGET_GROUPS.items():
        detail = build_group_detail(
            group_name, leader, intent_data["rows"],
            quality_data["lp_problems"], quality_data["lp_total"],
        )
        messages.append((f"{group_name} 跟进清单", detail))

    if args.test:
        for i, (title, content) in enumerate(messages, 1):
            if args.msg > 0 and i != args.msg:
                continue
            print(f"\n{'='*50}")
            print(f"消息 {i}: {title}")
            print(f"{'='*50}")
            print(content)
        return

    keys = args.key if isinstance(args.key, list) else [args.key]
    all_results = []

    for key_idx, key in enumerate(keys):
        if len(keys) > 1:
            print(f"\n--- 推送到群 {key_idx+1}/{len(keys)} ---")
        pusher = DingTalkPusher(key)
        results = []

        for i, (title, content) in enumerate(messages, 1):
            if args.msg > 0 and i != args.msg:
                continue
            # Trim if over 4096 chars
            if len(content) > 4096:
                content = content[:4050] + "\n> ... (内容超长已截断)"
            result = pusher.send_markdown(title, content)
            results.append({"msg": i, "title": title, **result})
            prefix = f"[群{key_idx+1}]" if len(keys) > 1 else ""
            print(f"{prefix}[{i}] {title}: errcode={result.get('errcode')}")
            if i < len(messages):
                time.sleep(1)

        all_results.extend(results)
        if key_idx < len(keys) - 1:
            time.sleep(1)

    success = all(r.get("errcode") == 0 for r in all_results)
    print(json.dumps({
        "status": "ok" if success else "partial_error",
        "segments": len(all_results),
        "keys": len(keys),
    }, ensure_ascii=False))
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()
