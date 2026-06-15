#!/usr/bin/env python3
"""续费意向日报：比较昨天 vs 今天的数据，推送意向等级变化。

用法:
  python daily_report.py [--today YYYY-MM-DD] [--test] [--key WEBHOOK_KEY]
"""
import argparse
import json
import sys
import time
import urllib.request
from pathlib import Path
from datetime import date, timedelta
from collections import Counter

import openpyxl

import config

OUTPUT_BASE = config.path("output_base")
TASK_NAME = config.get("task_name")
DEFAULT_KEYS = config.get("dingtalk_webhook_keys")


def read_intent(path: str) -> dict[str, dict]:
    """读取续费意向 Excel，返回 {sid: {level, score, lp, pos_signals, risk_signals}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    
    def exact_or_default(keyword: str, default: int, exclude_keywords: list = None) -> int:
        """精确匹配列名，排除含其他关键词的列(如'正向信号数'不应匹配'正向信号')"""
        for i, h in enumerate(headers):
            if keyword not in h:
                continue
            if exclude_keywords:
                if any(ex in h for ex in exclude_keywords):
                    continue
            return i
        return default
    
    col_sid = exact_or_default("学员ID", 1)
    col_level = exact_or_default("意向等级", 5)
    col_score = exact_or_default("综合得分", 6)
    col_lp = exact_or_default("归属LP", 2)
    col_pos = exact_or_default("正向信号", 10, exclude_keywords=["正向信号数", "正向信号计"])
    col_risk = exact_or_default("风险信号", 11, exclude_keywords=["风险信号数", "风险信号计"])
    
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[col_sid]).strip().rstrip(".0") if row[col_sid] else ""
        if not sid or sid == "None":
            continue
        level_raw = str(row[col_level]).strip() if row[col_level] else "LOW"
        level = "HIGH" if "HIGH" in level_raw or "高" in level_raw else \
                "MED" if "MED" in level_raw or "中" in level_raw else \
                "RISK" if "RISK" in level_raw or "风险" in level_raw else "LOW"
        result[sid] = {
            "level": level,
            "score": int(float(row[col_score])) if row[col_score] else 0,
            "lp": str(row[col_lp]).strip() if row[col_lp] else "",
            "pos_signals": str(row[col_pos]).strip() if row[col_pos] else "",
            "risk_signals": str(row[col_risk]).strip() if row[col_risk] else "",
        }
    wb.close()
    return result


def compare(yesterday: dict, today: dict) -> dict:
    """比较两天数据，找出意向等级变化"""
    level_order = {"HIGH": 4, "MED": 3, "LOW": 2, "RISK": 0}
    
    warming = []   # 升温: level up
    cooling = []   # 降温: level down
    new_risk = []  # 新增风险
    new_high = []  # 新增高意向 (包括昨天不存在的新学员)
    resolved_risk = []  # 风险已解除
    
    all_sids = set(yesterday.keys()) | set(today.keys())
    new_students = set(today.keys()) - set(yesterday.keys())
    gone_students = set(yesterday.keys()) - set(today.keys())
    
    for sid in all_sids:
        yd = yesterday.get(sid)
        td = today.get(sid)
        
        if yd is None:
            # 新学员
            if td["level"] == "HIGH":
                new_high.append({"sid": sid, "today": td})
            elif td["level"] == "RISK":
                new_risk.append({"sid": sid, "today": td, "from": "NEW"})
            continue
        
        if td is None:
            # 学员消失(可能是已签单被过滤等)
            continue
        
        old_lv = yd["level"]
        new_lv = td["level"]
        old_score = level_order.get(old_lv, 2)
        new_score = level_order.get(new_lv, 2)
        
        if new_score > old_score:
            warming.append({
                "sid": sid, "from": old_lv, "to": new_lv,
                "today": td, "yesterday": yd,
            })
        elif new_score < old_score:
            cooling.append({
                "sid": sid, "from": old_lv, "to": new_lv,
                "today": td, "yesterday": yd,
            })
        
        # 新增风险(之前不是RISK，现在是RISK)
        if old_lv != "RISK" and new_lv == "RISK":
            new_risk.append({"sid": sid, "today": td, "from": old_lv})
        
        # 风险已解除(之前是RISK，现在不是)
        if old_lv == "RISK" and new_lv != "RISK":
            resolved_risk.append({"sid": sid, "from": "RISK", "to": new_lv, "today": td})
    
    return {
        "warming": warming, "cooling": cooling,
        "new_risk": new_risk, "new_high": new_high,
        "resolved_risk": resolved_risk,
        "new_students": len(new_students),
        "gone_students": len(gone_students),
        "today_total": len(today),
        "yesterday_total": len(yesterday),
    }


def count_levels(data: dict) -> dict:
    c = Counter(d["level"] for d in data.values())
    return {"HIGH": c["HIGH"], "MED": c["MED"], "LOW": c["LOW"], "RISK": c["RISK"]}


def build_daily_brief(prev_date: str, curr_date: str, changes: dict, today_counts: dict, yesterday_counts: dict, feishu_url: str = "") -> str:
    """构建日报 Markdown"""
    lines = []
    lines.append(f"# 📊 续费意向日报")
    lines.append(f"### {prev_date} → {curr_date}")
    lines.append("")
    
    # ── 变化总览 ──
    n_warm = len(changes["warming"])
    n_cool = len(changes["cooling"])
    n_risk = len(changes["new_risk"])
    n_high = len(changes["new_high"])
    n_resolved = len(changes["resolved_risk"])
    n_new = changes["new_students"]
    n_gone = changes["gone_students"]
    
    lines.append("## ⚡ 昨日变化")
    lines.append("")
    
    parts = []
    if n_high > 0:
        parts.append(f"🆕 新高意向 {n_high}人")
    if n_warm > 0:
        parts.append(f"🔥 升温 {n_warm}人")
    if n_cool > 0:
        parts.append(f"❄️ 降温 {n_cool}人")
    if n_risk > 0:
        parts.append(f"🔴 新增风险 {n_risk}人")
    if n_resolved > 0:
        parts.append(f"✅ 风险解除 {n_resolved}人")
    
    if not parts:
        lines.append("> 昨日无意向等级变化")
        lines.append("")
    else:
        lines.append("> " + "　".join(parts))
        lines.append("")
    
    # ── 当前分布 ──
    lines.append("## 📈 当前意向分布")
    lines.append("")
    today_t = changes["today_total"]
    yd_t = changes["yesterday_total"]
    
    def delta_str(now, before):
        d = now - before
        if d > 0:
            return f"+{d}"
        return str(d)
    
    lines.append(f"| 等级 | 昨日 | 今日 | 变化 | 占比 |")
    lines.append(f"|------|:---:|:---:|:---:|:---:|")
    for lv in ["HIGH", "MED", "LOW", "RISK"]:
        yd_c = yesterday_counts[lv]
        td_c = today_counts[lv]
        d = delta_str(td_c, yd_c)
        pct = f"{td_c/today_t*100:.1f}%"
        lines.append(f"| {'🔥' if lv=='HIGH' else '🔴' if lv=='RISK' else '🟡' if lv=='MED' else '🔵'} {lv} | {yd_c} | {td_c} | {d} | {pct} |")
    lines.append("")
    
    # ── 升温明细(最多10人) ──
    all_warming = changes["new_high"] + changes["warming"]
    if all_warming:
        warming_sorted = sorted(all_warming, key=lambda x: -x["today"]["score"])[:10]
        lines.append("## 🔥 升温学员 (48h跟进)")
        lines.append("")
        lines.append(f"| 学员ID | LP | 变化 | 分 | 关键信号 |")
        lines.append(f"|--------|-----|------|:--:|---------|")
        for w in warming_sorted:
            td = w["today"]
            sid = td.get("sid", w["sid"])
            change = f"🆕" if "from" not in w or w.get("from") == "NEW" else f"{w['from']}→{w['to']}"
            sig = td["pos_signals"][:30] if td["pos_signals"] else "—"
            lines.append(f"| {sid} | {td['lp']} | {change} | {td['score']} | {sig} |")
        lines.append("")
    
    # ── 新增风险(最多10人) ──
    if changes["new_risk"]:
        risk_sorted = sorted(changes["new_risk"], key=lambda x: x["today"]["score"])[:10]
        lines.append("## 🔴 新增风险 (主管挽留)")
        lines.append("")
        lines.append(f"| 学员ID | LP | 来源 | 风险信号 |")
        lines.append(f"|--------|-----|------|---------|")
        for r in risk_sorted:
            td = r["today"]
            sig = td["risk_signals"][:40] if td["risk_signals"] else "—"
            lines.append(f"| {r['sid']} | {td['lp']} | {r['from']}→RISK | {sig} |")
        lines.append("")
    
    # ── 降温提醒 ──
    if changes["cooling"]:
        cooling_display = changes["cooling"][:8]
        lines.append("## ❄️ 降温提醒")
        lines.append("")
        lines.append(f"| 学员ID | LP | 变化 |")
        lines.append(f"|--------|-----|------|")
        for c in cooling_display:
            lines.append(f"| {c['sid']} | {c['today']['lp']} | {c['from']}→{c['to']} |")
        lines.append("")
    
    # ── 数据备注 ──
    lines.append("---")
    lines.append(f"数据: {changes['today_total']} 学员　　新入库 {n_new}人　　退库 {n_gone}人")
    if feishu_url:
        lines.append(f"📊 [意向客户明细表（飞书表格）]({feishu_url})")
    lines.append("")
    
    return "\n".join(lines)


def send_dingtalk(key: str, title: str, content: str, test: bool = False) -> dict:
    if test:
        print(f"\n{'='*60}")
        print(f"[TEST] {title}")
        print(f"{'='*60}")
        sys.stdout.buffer.write((content[:2000] + '\n').encode('utf-8'))
        sys.stdout.flush()
        return {"errcode": 0}
    
    url = f"https://oapi.dingtalk.com/robot/send?access_token={key}"
    data = json.dumps({
        "msgtype": "markdown",
        "markdown": {"title": title, "text": content}
    }, ensure_ascii=False)
    req = urllib.request.Request(
        url, data=data.encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8"},
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        body = resp.read().decode("utf-8")
    return json.loads(body)


def send_dingtalk_all(keys: list, title: str, content: str, test: bool = False):
    """向所有 webhook 群推送，返回所有结果"""
    results = []
    keys_list = keys if isinstance(keys, list) else [keys]
    for i, key in enumerate(keys_list):
        if len(keys_list) > 1:
            print(f"  群 {i+1}/{len(keys_list)}", end=" ")
        r = send_dingtalk(key, title, content, test=test)
        results.append(r)
        if not test:
            print(f"errcode={r.get('errcode')}", end="  ")
    if not test:
        print()
    return results


def main():
    parser = argparse.ArgumentParser(description="续费意向日报生成与推送")
    parser.add_argument("--today", help="今天的日期 (YYYY-MM-DD), 默认今天")
    parser.add_argument("--yesterday", help="昨天的日期 (YYYY-MM-DD), 默认昨天")
    parser.add_argument("--key", nargs="*", default=DEFAULT_KEYS,
                        help="钉钉 Webhook Key（可多个），默认从 config 读取全部")
    parser.add_argument("--test", action="store_true", help="测试模式: 打印不发送")
    parser.add_argument("--silent-on-no-change", action="store_true", help="无变化时不推送")
    parser.add_argument("--feishu-url", default="", help="飞书表格链接")
    args = parser.parse_args()
    
    today_date = date.fromisoformat(args.today) if args.today else date.today()
    yesterday_date = date.fromisoformat(args.yesterday) if args.yesterday else today_date - timedelta(days=1)
    
    # 全局搜索: 文件名匹配日期，不限目录（因为 run_date ≠ analysis_date）
    base = OUTPUT_BASE / TASK_NAME
    def find_intent_file(date_str: str) -> Path | None:
        candidates = list(base.rglob(f"renewal_intention_{date_str}.xlsx"))
        if candidates:
            return candidates[-1]  # 最新的
        candidates = list(base.rglob(f"renewal_intention_{date_str}*.xlsx"))
        return candidates[-1] if candidates else None
    
    today_file = find_intent_file(today_date.isoformat())
    yesterday_file = find_intent_file(yesterday_date.isoformat())
    
    if not today_file:
        print(f"ERROR: 找不到 {today_date} 的分析文件", file=sys.stderr)
        sys.exit(1)
    if not yesterday_file:
        # 首日运行：无昨日对比数据，发送当前分布简报
        if args.silent_on_no_change:
            print("首日运行，无对比数据，跳过推送")
            return
        print(f"首日运行: 无 {yesterday_date} 对比数据，发送当前分布简报")
        today_data = read_intent(str(today_file))
        today_counts = count_levels(today_data)
        first_day_brief = f"""# 📊 续费意向首日简报
### {today_date.isoformat()}

## 📈 当前意向分布

> 🔥 HIGH {today_counts['HIGH']}人　🟡 MED {today_counts['MED']}人　🔵 LOW {today_counts['LOW']}人　🔴 RISK {today_counts['RISK']}人

| 等级 | 人数 | 占比 |
|------|:---:|:---:|
| 🔥 HIGH | {today_counts['HIGH']} | {today_counts['HIGH']/len(today_data)*100:.1f}% |
| 🟡 MED | {today_counts['MED']} | {today_counts['MED']/len(today_data)*100:.1f}% |
| 🔵 LOW | {today_counts['LOW']} | {today_counts['LOW']/len(today_data)*100:.1f}% |
| 🔴 RISK | {today_counts['RISK']} | {today_counts['RISK']/len(today_data)*100:.1f}% |

---
📊 首日运行，明日开始将推送意向等级变化。
数据: {len(today_data)} 学员"""
        if args.feishu_url:
            first_day_brief += f"\n📊 [意向客户明细表（飞书表格）]({args.feishu_url})"
        
        brief = first_day_brief
        if args.test:
            send_dingtalk_all(args.key, f"续费意向首日简报 {today_date.isoformat()}", brief, test=True)
            return
        # 真实推送
        if len(brief) > 4096:
            brief = brief[:4050] + "\n> ..."
        title = f"续费意向首日简报 {today_date.isoformat()}"
        result = send_dingtalk_all(args.key, title, brief)
        errcodes = [r.get('errcode') for r in result]
        print(f"推送结果: errcode={errcodes}")
        return
    
    print(f"今天: {today_file}")
    print(f"昨天: {yesterday_file}")
    
    today_data = read_intent(str(today_file))
    yesterday_data = read_intent(str(yesterday_file))
    
    print(f"今天 {len(today_data)} 学员, 昨天 {len(yesterday_data)} 学员")
    
    changes = compare(yesterday_data, today_data)
    today_counts = count_levels(today_data)
    yesterday_counts = count_levels(yesterday_data)
    
    total_changes = len(changes["warming"]) + len(changes["cooling"]) + \
                    len(changes["new_risk"]) + len(changes["new_high"]) + \
                    len(changes["resolved_risk"])
    
    if total_changes == 0 and args.silent_on_no_change:
        print("无变化，跳过推送")
        return
    
    brief = build_daily_brief(
        yesterday_date.isoformat(), today_date.isoformat(),
        changes, today_counts, yesterday_counts,
        args.feishu_url,
    )
    
    # 推送
    if len(brief) > 4096:
        brief = brief[:4050] + "\n> ... (内容超长已截断)"
    
    title = f"续费意向日报 {today_date.isoformat()}"
    result = send_dingtalk_all(args.key, title, brief, test=args.test)
    
    if args.test:
        print(f"\n[{len(brief)} chars]")
    else:
        errcodes = [r.get('errcode') for r in result]
        print(f"推送结果: errcode={errcodes}")
        all_ok = all(ec == 0 for ec in errcodes)
        if not all_ok:
            print(f"ERROR: {result}", file=sys.stderr)
            sys.exit(1)


if __name__ == "__main__":
    main()
