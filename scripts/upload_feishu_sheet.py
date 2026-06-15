#!/usr/bin/env python3
"""续费意向 Excel → 飞书表格（使用 lark-cli，分批避开命令行长度限制）

用法:
  python upload_feishu_sheet.py <意向Excel路径> [--title 标题]
"""
import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from datetime import date, datetime
import tempfile, os

import openpyxl

import config

LARK_CLI = str(config.path("lark_cli"))
BATCH = 100  # 每批行数，太小Json不会超命令行


def run_lark(args: list[str]) -> dict:
    """运行 lark-cli，返回解析的 JSON"""
    cmd = [LARK_CLI] + args + ["--as", "user"]
    result = subprocess.run(
        cmd, capture_output=True, text=True, encoding="utf-8", timeout=120
    )
    if result.returncode != 0:
        err = result.stderr or result.stdout
        raise RuntimeError(f"lark-cli error (rc={result.returncode}): {err[:300]}")
    try:
        return json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        return {"_raw": result.stdout}


def read_excel(path: str) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals, has = [], False
        for c in row:
            if c is None:
                vals.append("")
            elif isinstance(c, float) and c == int(c):
                vals.append(str(int(c)))
            else:
                s = str(c).strip()
                vals.append(s)
                if s:
                    has = True
        if has:
            rows.append(vals)
    wb.close()
    return headers, rows


def read_raw_bio_lookup(raw_path: str) -> dict:
    """从原始 BI 导出建立 {学员ID: {小组, 池子, 续费金额, 语种}} 映射"""
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Find header row (usually row 6, 1-indexed)
    header_row = None
    for r in range(1, min(12, ws.max_row + 1)):
        row_vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
        if "学员ID" in row_vals and "小组" in row_vals:
            header_row = r
            break
    
    if not header_row:
        wb.close()
        return {}
    
    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
    sid_col = next((i for i, h in enumerate(headers) if h == "学员ID"), 0)
    group_col = next((i for i, h in enumerate(headers) if h == "小组"), 3)
    pool_col = next((i for i, h in enumerate(headers) if h == "池子"), None)
    amount_col = next((i for i, h in enumerate(headers) if h == "续费金额"), None)
    lang_col = next((i for i, h in enumerate(headers) if h == "语种"), None)
    contact_col = next((i for i, h in enumerate(headers) if h == "沟通时间"), None)
    exec_col = next((i for i, h in enumerate(headers) if h == "是否执行"), None)
    
    lookup = {}
    today = date.today()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sid = str(row[sid_col]).strip().rstrip(".0") if row[sid_col] else ""
        if not sid or sid == "None":
            continue
        if sid not in lookup:
            entry = {}
            if group_col is not None and row[group_col]:
                entry["小组"] = str(row[group_col]).strip()
            if pool_col is not None and row[pool_col]:
                entry["池子"] = str(row[pool_col]).strip()
            if amount_col is not None and row[amount_col]:
                entry["续费金额"] = str(row[amount_col]).strip()
            if lang_col is not None and row[lang_col]:
                entry["语种"] = str(row[lang_col]).strip()
            if entry:
                lookup[sid] = entry
            # Track last contact time per student
            if contact_col is not None:
                contact_val = row[contact_col]
                if contact_val:
                    try:
                        if isinstance(contact_val, str):
                            ct = datetime.strptime(contact_val[:10], "%Y-%m-%d").date()
                        elif hasattr(contact_val, 'date'):
                            ct = contact_val.date() if hasattr(contact_val, 'date') else contact_val
                        else:
                            ct = None
                        if ct:
                            days_gap = (today - ct).days
                            entry["失联天数"] = str(days_gap)
                            entry["是否失联超过7天"] = "是" if days_gap > 7 else "否"
                    except:
                        pass
            # Track action count (temp storage via dict)
            if exec_col is not None and row[exec_col] == 1.0:
                entry["_action_count"] = entry.get("_action_count", 0) + 1
    
    # Convert action counts + default missing fields
    for sid in lookup:
        ac = lookup[sid].pop("_action_count", 0)
        lookup[sid]["近7天动作数"] = str(ac)
        if "是否失联超过7天" not in lookup[sid]:
            lookup[sid]["是否失联超过7天"] = "无数据"
        # 断联 = 曾被覆盖(有动作) 且 失联>7天
        lost_flag = lookup[sid].get("是否失联超过7天", "否")
        has_act = ac > 0
        lookup[sid]["是否断联"] = "是" if (lost_flag == "是" and has_act) else "否"
    
    wb.close()
    return lookup


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_path")
    parser.add_argument("--title", default=None)
    parser.add_argument("--raw-excel", default=None, help="原始 BI 导出路径，用于补充小组/池子字段")
    parser.add_argument("--process-excel", default=None, help="过程报告 Excel 路径，上传为第二个 Sheet")
    parser.add_argument("--low-excel", default=None, help="低意向分析 Excel 路径，上传为第三个 Sheet")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    
    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        print(f"ERROR: 文件不存在: {excel_path}", file=sys.stderr)
        sys.exit(1)
    
    today = date.today()
    title = args.title or f"续费意向盘 {today.isoformat()}"
    
    if not args.json:
        print(f"读取: {excel_path.name}")
    
    headers, rows = read_excel(str(excel_path))
    
    # Merge raw BI data: fill existing empty columns, add new columns
    if args.raw_excel and Path(args.raw_excel).exists():
        if not args.json:
            print(f"补充字段: {args.raw_excel}")
        lookup = read_raw_bio_lookup(args.raw_excel)
        
        # Map: extra_field → header name in raw BI
        # "小组" exists in analysis (col 3) but is empty → fill it
        # "池子", "续费金额", "语种" are new → add as new columns
        sid_idx = next((i for i, h in enumerate(headers) if h == "学员ID"), 1)
        group_idx = next((i for i, h in enumerate(headers) if h == "小组"), None)
        
        # New columns to add (after "区域")
        new_fields = ["池子", "续费金额", "语种", "失联天数", "近7天动作数", "是否失联超过7天", "是否断联"]
        insert_pos = next((i for i, h in enumerate(headers) if h == "区域"), 
                          group_idx if group_idx is not None else 3)
        # Insert headers in reverse so final order matches
        for f in reversed(new_fields):
            headers.insert(insert_pos + 1, f)
        
        for row in rows:
            sid = str(row[sid_idx]).strip().rstrip(".0") if len(row) > sid_idx else ""
            info = lookup.get(sid, {})
            # Fill existing "小组" column if empty
            if group_idx is not None and group_idx < len(row):
                if not row[group_idx] and info.get("小组"):
                    row[group_idx] = info["小组"]
            # Fill "区域" from "小组" (derive: remove trailing number/group suffix)
            region_idx = next((i for i, h in enumerate(headers) if h == "区域"), None)
            if region_idx is not None and region_idx < len(row):
                group_val = row[group_idx] if group_idx is not None and group_idx < len(row) else ""
                if group_val and not row[region_idx]:
                    # 港澳1组→港澳, 美澳2组→美澳, 台湾组→台湾
                    region = re.sub(r'\d+组$', '', group_val)
                    region = re.sub(r'组$', '', region)
                    row[region_idx] = region
            # Insert row values: same order as headers (reversed so final order matches headers)
            pos = insert_pos + 1
            for f in reversed(new_fields):
                row.insert(pos, info.get(f, ""))
    
    short_headers = [h[:20] for h in headers]
    
    if not args.json:
        print(f"数据: {len(rows)} 行 × {len(headers)} 列")
    
    # Step 1: Create spreadsheet with headers only (small)
    if not args.json:
        print("创建飞书表格...")
    
    r = run_lark([
        "sheets", "+create",
        "--title", title,
        "--headers", json.dumps(short_headers, ensure_ascii=False),
        "--data", json.dumps([], ensure_ascii=False),
    ])
    
    token = r.get("spreadsheetToken") or \
            (r.get("data", {}) or {}).get("spreadsheet_token") or \
            (r.get("data", {}) or {}).get("spreadsheet", {}).get("spreadsheet_token") or \
            (r.get("data", {}) or {}).get("spreadsheet", {}).get("spreadsheetToken")
    
    # Extract sheet_id from create response or fetch it
    sheet_id = None
    data = r.get("data", {}) or {}
    sheets_list = data.get("sheets") or \
                  data.get("spreadsheet", {}).get("sheets") or []
    if sheets_list:
        sheet_id = sheets_list[0].get("sheet_id") or sheets_list[0].get("sheetId")
    
    if not sheet_id:
        print("  (使用默认第一个 Sheet)")
    url = r.get("url") or \
          (r.get("data", {}) or {}).get("url") or \
          (r.get("data", {}) or {}).get("spreadsheet", {}).get("url")
    
    if not token:
        print(f"ERROR: 无法获取 token from: {json.dumps(r, ensure_ascii=False)[:500]}", file=sys.stderr)
        sys.exit(1)
    
    if not args.json:
        print(f"  表格创建: {token}, sheet={sheet_id}")
    
    # Step 2: Append data in batches (small enough for command line)
    total = len(rows)
    for i in range(0, total, BATCH):
        batch = rows[i:i + BATCH]
        end = min(i + BATCH, total)
        
        if not args.json:
            print(f"  写入 {i+1}-{end}/{total} 行...")
        
        append_args = [
            "sheets", "+append",
            "--spreadsheet-token", token,
            "--values", json.dumps(batch, ensure_ascii=False),
        ]
        if sheet_id:
            append_args.extend(["--sheet-id", sheet_id])
        
        run_lark(append_args)
    
    # Step 2.5: Upload process report as Sheet2
    if args.process_excel and Path(args.process_excel).exists():
        if not args.json:
            print("上传过程分析表 (Sheet2)...")
        
        # Read process report data
        pw = openpyxl.load_workbook(args.process_excel, data_only=True)
        pws = pw.active
        process_headers = [str(pws.cell(row=2, column=c).value or '')[:20].strip() 
                          if pws.cell(row=2, column=c).value else ''
                          for c in range(1, 15)]
        # Fill empty headers for cols 1-2 from merged cells
        process_headers[0] = '团队'
        process_headers[1] = '池子'
        # Fill cols 12-13 (merged cells)
        process_headers[11] = '断联率'
        process_headers[12] = '退费占比'
        process_headers[13] = '结课占比'
        
        process_rows = []
        for r in range(3, pws.max_row + 1):
            vals = [str(pws.cell(row=r, column=c).value or '').strip() for c in range(1, 15)]
            if any(v for v in vals[:3]):
                process_rows.append(vals)
        pw.close()
        
        # Create Sheet2
        r2 = run_lark([
            "sheets", "+create-sheet",
            "--spreadsheet-token", token,
            "--title", "续费过程分析",
            "--as", "user",
        ])
        sheet2_id = (r2.get("data", {}) or {}).get("sheet_id") or ""
        
        # Append process data
        batch2 = [process_headers] + process_rows
        for k in range(0, len(batch2), BATCH):
            chunk = batch2[k:k + BATCH]
            append_args2 = [
                "sheets", "+append",
                "--spreadsheet-token", token,
                "--values", json.dumps(chunk, ensure_ascii=False),
            ]
            if sheet2_id:
                append_args2.extend(["--sheet-id", sheet2_id])
            run_lark(append_args2)
        
        if not args.json:
            print(f"  过程分析表: {len(process_rows)} 行 × {len(process_headers)} 列")
    
    # Step 2.6: Upload low intention analysis as Sheet3
    if args.low_excel and Path(args.low_excel).exists():
        if not args.json:
            print("上传低意向分析 (Sheet3)...")
        lw = openpyxl.load_workbook(args.low_excel, data_only=True)
        lws = lw.active
        low_headers = [str(lws.cell(row=1, column=c).value or '') for c in range(1, 12)]
        low_rows = []
        for r in range(2, lws.max_row + 1):
            vals = [str(lws.cell(row=r, column=c).value or '') for c in range(1, 12)]
            if any(v for v in vals[:3]):
                low_rows.append(vals)
        lw.close()
        r3 = run_lark([
            "sheets", "+create-sheet",
            "--spreadsheet-token", token,
            "--title", "低意向原因分析",
            "--as", "user",
        ])
        sheet3_id = (r3.get("data", {}) or {}).get("sheet_id") or ""
        batch3 = [low_headers] + low_rows
        for k in range(0, len(batch3), BATCH):
            chunk = batch3[k:k + BATCH]
            a3 = [
                "sheets", "+append",
                "--spreadsheet-token", token,
                "--values", json.dumps(chunk, ensure_ascii=False),
            ]
            if sheet3_id:
                a3.extend(["--sheet-id", sheet3_id])
            run_lark(a3)
        if not args.json:
            print(f"  低意向分析: {len(low_rows)} 行 × {len(low_headers)} 列")
    
    # Step 3: Set public permission & get link
    if not args.json:
        print("设置分享链接...")
    
    perm_data = json.dumps({
        "link_share_entity": "tenant_readable",
        "share_entity": "anyone",
        "external_access": True,
        "invite_external": True,
    }, ensure_ascii=False)
    
    try:
        run_lark([
            "drive", "permission.public", "patch",
            token,  # path param: {token}/public
            "--data", perm_data,
            "--params", json.dumps({"type": "sheet"}),
            "--yes",
        ])
    except Exception as e:
        if not args.json:
            print(f"  权限设置跳过: {e}")
    
    share_url = url or f"https://hcnig43mb8gp.feishu.cn/sheets/{token}"
    
    output = {
        "status": "ok",
        "spreadsheet_token": token,
        "url": share_url,
        "title": title,
        "rows": len(rows),
        "columns": len(headers),
    }
    
    if args.json:
        print(json.dumps(output, ensure_ascii=False))
    else:
        print(f"\n✓ 飞书表格已创建")
        print(f"  链接: {share_url}")
    
    return output


if __name__ == "__main__":
    main()
