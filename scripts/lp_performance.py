#!/usr/bin/env python3
"""LP 续费绩效卡: 每人跟进覆盖率/及时率/破零/签单/模板依赖

用法:
  python lp_performance.py <原始BI.xlsx> [--intent <意向.xlsx>] [--output-dir .]
"""
import argparse, os, sys, re
from collections import defaultdict, Counter
from datetime import date, datetime
import openpyxl


def main():
    parser = argparse.ArgumentParser(description="LP续费绩效卡")
    parser.add_argument("raw_excel")
    parser.add_argument("--intent", default="", help="意向分析 Excel")
    parser.add_argument("--output-dir", default=".")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.raw_excel, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, min(12, ws.max_row + 1)):
        row_vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
        if "学员ID" in row_vals and "当前LP" in row_vals:
            header_row = r
            break

    if not header_row:
        print("ERROR: 找不到表头", file=sys.stderr)
        sys.exit(1)

    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
    sid_col = next((i for i, h in enumerate(headers) if h == "学员ID"), 0)
    lp_col = next((i for i, h in enumerate(headers) if h == "当前LP"), 2)
    group_col = next((i for i, h in enumerate(headers) if h == "小组"), 3)
    scene_col = next((i for i, h in enumerate(headers) if h == "场景名称"), 4)
    sem_col = next((i for i, h in enumerate(headers) if h == "语义点"), 5)
    exec_col = next((i for i, h in enumerate(headers) if h == "是否执行"), 6)
    result_col = next((i for i, h in enumerate(headers) if h == "执行结果"), 7)
    signed_col = next((i for i, h in enumerate(headers) if "是否续费" in h), 8)
    contact_col = next((i for i, h in enumerate(headers) if h == "沟通时间"), 18)
    dialog_col = next((i for i, h in enumerate(headers) if "材料全对话" in h), 20)

    today = date.today()
    lp_data = defaultdict(lambda: {
        "小组": "", "学员数": 0, "已签": 0, "未签": 0,
        "外呼数": 0, "跟进数": 0, "失联>7天": 0, "模板数": 0,
        "学情反馈数": 0, "退费保障数": 0, "续费话术数": 0,
    })

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        lp = str(row[lp_col]).strip() if row[lp_col] else ""
        if not lp or lp == "None":
            continue
        d = lp_data[lp]
        if not d["小组"]:
            d["小组"] = str(row[group_col]).strip() if row[group_col] else ""

        sid = str(row[sid_col]).strip().rstrip(".0") if row[sid_col] else ""
        scene = str(row[scene_col]).strip() if row[scene_col] else ""
        sem = str(row[sem_col]).strip() if row[sem_col] else ""
        executed = row[exec_col]
        result = str(row[result_col]).strip() if row[result_col] else ""

        # Per-row: count actions
        if executed == 1.0:
            d["跟进数"] += 1

        # Detect template messages
        dialog = str(row[dialog_col]).strip() if row[dialog_col] else ""
        if len(dialog) > 50:
            # Template: repeated identical prefix > 3 times
            lines = dialog.split("\n")
            prefix_count = Counter(l[:15] for l in lines if len(l) >= 15)
            if max(prefix_count.values(), default=0) >= 5:
                d["模板数"] += 1

        # Specific action types
        if "学情反馈" in sem or "学习规划" in sem:
            d["学情反馈数"] += 1
        if "退费保障" in result or "退费" in sem:
            d["退费保障数"] += 1
        if "续费结果" in sem or "续费" in scene:
            d["续费话术数"] += 1

        # Check contact recency
        if contact_col is not None and row[contact_col]:
            try:
                ct = str(row[contact_col])[:10]
                ct_date = datetime.strptime(ct, "%Y-%m-%d").date()
                if (today - ct_date).days > 7:
                    d["失联>7天"] += 1
            except:
                pass

    wb.close()

    # Per-student aggregation
    students_per_lp = defaultdict(set)
    signed_per_lp = defaultdict(set)
    wb2 = openpyxl.load_workbook(args.raw_excel, data_only=True)
    ws2 = wb2.active
    for row in ws2.iter_rows(min_row=header_row + 1, values_only=True):
        sid = str(row[sid_col]).strip().rstrip(".0") if row[sid_col] else ""
        lp = str(row[lp_col]).strip() if row[lp_col] else ""
        if not sid or not lp:
            continue
        students_per_lp[lp].add(sid)
        if row[signed_col] == 1.0:
            signed_per_lp[lp].add(sid)
    wb2.close()

    # Write output
    out_wb = openpyxl.Workbook()
    ws_out = out_wb.active
    ws_out.title = "LP绩效卡"
    hfill = openpyxl.styles.PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfont = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)

    hdrs = ["LP", "小组", "学员数", "已签", "签单率", "跟进数", "人均跟进", "失联>7天",
            "模板数", "学情反馈", "退费保障", "续费话术"]
    for ci, h in enumerate(hdrs, 1):
        c = ws_out.cell(row=1, column=ci, value=h)
        c.font, c.fill = hfont, hfill

    lp_list = []
    for lp, d in lp_data.items():
        total_s = len(students_per_lp[lp])
        signed_s = len(signed_per_lp[lp])
        if total_s == 0:
            continue
        lp_list.append({
            "lp": lp, "group": d["小组"],
            "total": total_s, "signed": signed_s,
            "sign_rate": signed_s / total_s,
            "follows": d["跟进数"],
            "avg_follow": d["跟进数"] / total_s,
            "lost_7d": d["失联>7天"],
            "template": d["模板数"],
            "feedback": d["学情反馈数"],
            "refund_guard": d["退费保障数"],
            "renewal_script": d["续费话术数"],
        })

    lp_list.sort(key=lambda x: (-x["sign_rate"], -x["total"]))
    for ri, r in enumerate(lp_list, 2):
        ws_out.cell(row=ri, column=1, value=r["lp"])
        ws_out.cell(row=ri, column=2, value=r["group"])
        ws_out.cell(row=ri, column=3, value=r["total"])
        ws_out.cell(row=ri, column=4, value=r["signed"])
        ws_out.cell(row=ri, column=5, value=f"{r['sign_rate']*100:.1f}%")
        ws_out.cell(row=ri, column=6, value=r["follows"])
        ws_out.cell(row=ri, column=7, value=f"{r['avg_follow']:.1f}")
        ws_out.cell(row=ri, column=8, value=r["lost_7d"])
        ws_out.cell(row=ri, column=9, value=r["template"])
        ws_out.cell(row=ri, column=10, value=r["feedback"])
        ws_out.cell(row=ri, column=11, value=r["refund_guard"])
        ws_out.cell(row=ri, column=12, value=r["renewal_script"])

        # Color coding
        if r["lost_7d"] > 2:
            ws_out.cell(row=ri, column=8).fill = openpyxl.styles.PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        if r["avg_follow"] < 2.0:
            ws_out.cell(row=ri, column=7).fill = openpyxl.styles.PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for ci, w in enumerate([10, 10, 8, 6, 8, 8, 8, 10, 8, 8, 8, 8], 1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out_path = os.path.join(args.output_dir, f"lp_performance_{today}.xlsx")
    out_wb.save(out_path)

    print(f"LP绩效卡: {out_path}")
    print(f"总计: {len(lp_list)}位LP")
    top5 = [r for r in lp_list[:5] if r["signed"] > 0]
    if top5:
        print(f"签单Top: " + ", ".join(f"{r['lp']}({r['signed']})" for r in top5))
    lost_lps = [r for r in lp_list if r["lost_7d"] > r["total"] * 0.3]
    if lost_lps:
        print(f"失联率高: " + ", ".join(f"{r['lp']}({r['lost_7d']}/{r['total']})" for r in lost_lps[:5]))

    return lp_list


if __name__ == "__main__":
    main()
