#!/usr/bin/env python3
"""低意向客户原因分类统计: 区域×小组 维度下按4类原因做占比分析

用法:
  python low_intention_analysis.py <续费意向Excel> <原始BI导出Excel> [--output-dir .]
"""
import argparse, os, re
from collections import defaultdict, OrderedDict
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


REASON_TYPES = OrderedDict([
    ("无互动", "总消息数=0，家长无任何回复"),
    ("纯观望", "有互动但无明确信号，被动接收"),
    ("意愿不足", "有正向信号但得分<1.5，未达中意向"),
    ("负面拉低", "有风险信号或得分<0"),
])


def classify(row: dict) -> str:
    msg = row.get("msg", 0)
    score = row.get("score", 0)
    risk = row.get("risk", "")
    has_risk = bool(risk and risk != "—")

    if msg == 0:
        return "无互动"
    if has_risk or score < 0:
        return "负面拉低"
    if score > 0 and score < 1.5:
        return "意愿不足"
    return "纯观望"


def region_from_group(group: str) -> str:
    return re.sub(r'\d+组$', '', group).rstrip('组')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("intent_excel")
    parser.add_argument("raw_excel")
    parser.add_argument("--output-dir", default=".")
    args = parser.parse_args()

    today = date.today()

    # Read raw BI for 小组 mapping
    rwb = openpyxl.load_workbook(args.raw_excel, data_only=True)
    rws = rwb.active
    header_row = None
    for r in range(1, min(15, rws.max_row + 1)):
        rv = [str(c.value).strip() if c.value else "" for c in rws[r]]
        if "学员ID" in rv and "小组" in rv:
            header_row = r; break
    if not header_row:
        header_row = 6
    rhdrs = [str(c.value).strip() if c.value else "" for c in rws[header_row]]
    rsid_c = next(i for i, h in enumerate(rhdrs) if h == "学员ID")
    rgrp_c = next(i for i, h in enumerate(rhdrs) if h == "小组")
    sid_group = {}
    for row in rws.iter_rows(min_row=header_row + 1, values_only=True):
        sid = str(row[rsid_c]).strip().rstrip(".0") if row[rsid_c] else ""
        gp = str(row[rgrp_c]).strip() if row[rgrp_c] else ""
        if sid and sid != "None" and gp and gp != "None":
            sid_group[sid] = gp
    rwb.close()

    # Read intention data
    iwb = openpyxl.load_workbook(args.intent_excel, data_only=True)
    iws = iwb.active
    ihdrs = [str(c.value).strip() if c.value else "" for c in iws[1]]
    sid_c = ihdrs.index("学员ID")
    lv_c = ihdrs.index("意向等级")
    sc_c = ihdrs.index("综合得分")
    ps_c = ihdrs.index("正向信号")
    rs_c = ihdrs.index("风险信号")
    ms_c = ihdrs.index("总消息数")

    # Aggregate: region → group → reason → count
    stats = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for row in iws.iter_rows(min_row=2, values_only=True):
        lv = str(row[lv_c]).strip() if row[lv_c] else ""
        if "LOW" not in lv:
            continue
        sid = str(row[sid_c]).strip().rstrip(".0") if row[sid_c] else ""
        r = {
            "score": float(row[sc_c]) if row[sc_c] else 0,
            "risk": str(row[rs_c]).strip() if row[rs_c] else "",
            "msg": int(row[ms_c]) if row[ms_c] else 0,
        }
        reason = classify(r)
        group = sid_group.get(sid, sid_group.get(sid + ".0", "未知"))
        region = region_from_group(group)
        stats[region][group][reason] += 1

    iwb.close()

    # Output Excel
    out_path = os.path.join(args.output_dir, f"low_intention_{today}.xlsx")
    owb = openpyxl.Workbook()
    ows = owb.active
    ows.title = "低意向原因分析"

    hfill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    hfont = Font(bold=True, size=11)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    dalign = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    reasons_list = list(REASON_TYPES.keys())
    headers = ["区域", "小组", "低意向总人数"] + reasons_list + [f"{r}占比" for r in reasons_list]
    col_widths = [10, 13, 12] + [10] * len(reasons_list) + [11] * len(reasons_list)

    for ci, h in enumerate(headers, 1):
        c = ows.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, thin
    for ci, w in enumerate(col_widths, 1):
        ows.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ri = 2
    all_regions = sorted(stats.keys())
    merge_deferred = []  # (start_row, end_row) for region merge

    for region in all_regions:
        region_groups = stats[region]
        sorted_groups = sorted(region_groups.items(), key=lambda x: -sum(x[1].values()))
        n = len(sorted_groups)
        start_r = ri

        for gi, (group, counts) in enumerate(sorted_groups):
            total = sum(counts.values())
            vals = [region if gi == 0 else "", group, total]
            for r in reasons_list:
                vals.append(counts.get(r, 0))
            for r in reasons_list:
                cnt = counts.get(r, 0)
                vals.append(f"{cnt/max(total,1)*100:.1f}%")

            for ci, v in enumerate(vals, 1):
                c = ows.cell(row=ri, column=ci, value=v)
                c.border = thin
                c.alignment = dalign

            ri += 1

        if n > 1:
            merge_deferred.append((start_r, start_r + n - 1))
        ri += 1  # separator

    for start_r, end_r in merge_deferred:
        ows.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)

    ows.freeze_panes = "A2"
    owb.save(out_path)
    print(f"低意向分析: {out_path}")

    # Summary
    total_low = sum(sum(rc.values()) for reg in stats.values() for rc in reg.values())
    reason_totals = defaultdict(int)
    for reg in stats.values():
        for rc in reg.values():
            for rn, cnt in rc.items():
                reason_totals[rn] += cnt
    print(f"总计: {total_low} 人")
    for rn in reasons_list:
        cnt = reason_totals.get(rn, 0)
        print(f"  {rn}: {cnt} ({cnt/max(total_low,1)*100:.1f}%)")


if __name__ == "__main__":
    main()
