"""
VIPTHINK LP 续费意向盘 — 核心分析脚本
========================================
从 LP-家长对话 Excel 中提取续费意向信号，生成：
  1. 学员意向明细表（Excel，可筛选）
  2. 管理者分析报告（Markdown）

用法:
  python analyze_renewal.py <输入Excel路径> [--output-dir <输出目录>]

输入格式:
  Excel 文件，必须包含以下列：
    - 学员ID、LP、材料全对话、沟通时间
    可选：区域等级、语义点、过程id

输出:
  - renewal_intention_YYYY-MM-DD.xlsx  （学员明细表）
  - renewal_report_YYYY-MM-DD.md       （管理报告）
"""

import re
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 信号定义（与 references/signal_taxonomy.md 同步）
# ============================================================

@dataclass
class Signal:
    name: str
    keywords: list          # 触发关键词列表
    weight: int             # 权重（正=加分，负=减分）
    category: str           # 分类：price/planning/commit/negative/risk
    description: str

# 高意向信号
HIGH_SIGNALS = [
    Signal("主动询价", ["续费多少钱", "什么价格", "课包怎么买", "现在多少钱", "续费价格", "价格多少", "怎么收费", "课程费用", "费用多少", "什么价位"], 3, "price", "主动询问续费价格"),
    Signal("问具体方案", ["有哪些方案", "哪个划算", "推荐哪个", "方案对比", "有什么课包", "课包推荐", "怎么选"], 3, "price", "在选择课包方案"),
    Signal("问优惠活动", ["有优惠吗", "还有活动", "能参加吗", "有折扣", "有活动吗", "能便宜", "优惠还有吗"], 3, "price", "在等优惠时机"),
    Signal("问升舱期", ["升舱", "升舱期", "升舱优惠", "升舱活动", "补差价"], 3, "price", "关注升舱期"),
    Signal("价格敏感/比价", ["算下来", "平均一节", "每个月", "对比", "别的地方", "别的机构", "其他平台", "别家", "别家", "划算吗", "值不值", "合算"], 3, "price", "价格敏感——在认真算账"),
    Signal("讨价还价", ["能不能便宜", "有折扣吗", "可以优惠", "最低多少", "能不能少", "打折", "给个优惠"], 3, "price", "讨价还价"),
    Signal("问下级别内容", ["下个级别", "下一级", "s3学什么", "s4学什么", "s5学什么", "之后学什么", "后面学什么", "升级学什么"], 2, "planning", "在规划后续学习"),
    Signal("长期学习规划", ["学完要多久", "几年学完", "能学到几岁", "学到什么时候", "一直学下去", "长期学"], 2, "planning", "有长期学习预期"),
    Signal("进度确认", ["还剩多少", "学到哪", "多少课了", "上了几节", "课时还有多少", "剩余课时", "进度怎样"], 2, "planning", "在盘点课时"),
    Signal("问付款方式", ["怎么付款", "怎么支付", "支持什么支付", "怎么交钱", "付款方式", "转账", "微信支付"], 2, "commit", "已进入付款阶段"),
    Signal("问冻结规则", ["冻结", "什么时候开始", "旧课包没", "新课包开启", "课包冻结", "先买放着"], 2, "commit", "在算时间窗口"),
    Signal("保价确认", ["涨价", "之后会涨", "锁定价格", "保价", "6月涨价", "v9涨价", "贵了"], 2, "commit", "对涨价有紧迫感"),
    Signal("明确表态续费", ["会续的", "继续学", "要续", "肯定续", "续费", "接着上", "会接着"], 2, "commit", "正面承诺续费"),
    Signal("主动问续费", ["续费怎么", "续课", "再买", "再报", "继续报", "续约"], 3, "commit", "主动发起续费话题"),
]

# 中意向信号
MEDIUM_SIGNALS = [
    Signal("正面回应续费", ["好的知道了", "谢谢提醒", "收到", "明白", "了解", "ok", "好的", "好滴", "好哒"], 1, "neutral", "不抵触但未推进"),
    Signal("对涨价有反应", ["要涨价", "涨价了", "现在买划算", "现在买便宜"], 1, "neutral", "感知涨价信息"),
    Signal("认可课程价值", ["确实进步", "上课还不错", "学得不错", "有进步", "喜欢上课", "效果挺好", "孩子喜欢"], 1, "neutral", "满意度OK"),
    Signal("需要商量", ["商量", "问下爸爸", "问下妈妈", "讨论一下", "家人商量", "和老公", "和老婆"], 1, "neutral", "多人决策"),
    Signal("表示再考虑", ["再考虑", "考虑一下", "想想", "想一下", "考虑下"], 1, "neutral", "未下决心"),
    Signal("时间不合适", ["现在太早", "课时还多", "等用完", "等上完", "后面再", "到时候"], 1, "neutral", "窗口未到"),
]

# 观望信号（权重0，仅标记）
LOW_SIGNALS = [
    Signal("敷衍回应", ["嗯", "好", "ok", "哦", "额"], 0, "silence", "极简回复"),
    Signal("模糊表态", ["到时候再说", "再看看", "看看吧", "不一定", "可能不"], 0, "silence", "刻意回避"),
]

# 风险信号
RISK_SIGNALS = [
    Signal("明确拒绝续费", ["不续了", "不学了", "不买了", "不续费", "不打算续", "不想续", "停止上课", "不想学"], -4, "risk", "直接表态不续"),
    Signal("课程不满", ["没效果", "学不到", "太简单", "太难", "跟不上", "不适合", "没什么用", "作用不大", "没进步"], -3, "risk", "价值感低"),
    Signal("替代方案", ["报了别的", "别的班", "换了机构", "没时间学", "要上学了", "学校作业多", "时间冲突"], -3, "risk", "已有替代方案"),
    Signal("投诉老师", ["老师不行", "老师不专业", "老师不好", "老师差", "换老师", "不喜欢老师"], -3, "risk", "对老师不满"),
    Signal("投诉系统", ["出问题", "老卡", "总闪退", "app不行", "系统太差", "进不去"], -3, "risk", "体验差"),
    Signal("服务不满", ["服务太差", "没人管", "不理人", "找不到人", "不理我"], -3, "risk", "对服务不满"),
    Signal("询问退款", ["怎么退款", "能退多少", "不上了退", "退费怎么", "退款流程", "退款要多久", "退钱", "可以退吗", "退款政策", "退课"], -4, "risk", "已经动退款念头"),
    Signal("已发起退款", ["我要退款", "帮我退", "我要退", "申请退款", "提交退款"], -4, "risk", "退款执行中"),
    Signal("长期失联", [], -2, "risk", "30天+无回复（时间维度判断）"),
]

# 所有信号合并
ALL_SIGNALS = HIGH_SIGNALS + MEDIUM_SIGNALS + LOW_SIGNALS + RISK_SIGNALS

# 等级阈值（方案E：分类上限法）
LEVEL_THRESHOLDS = {
    "HIGH": 3,      # ≥3
    "MEDIUM": 1.5,  # 1.5~2.9
    "RISK": -1,     # ≤-1
}


# ============================================================
# 对话解析
# ============================================================

def parse_conversations(raw_text: str) -> list[dict]:
    """
    解析「材料全对话」字段。
    格式：「员工:xxx」/「客户:xxx」交替出现。
    返回：[{"role": "client"/"lp", "content": "..."}, ...]
    """
    if not raw_text or not isinstance(raw_text, str):
        return []

    raw_text = raw_text.strip()
    lines = []
    # 尝试按"员工:"、"客户:"、"家长:"、"班主任:"、"老师:"等角色标记分割
    parts = re.split(r'(员工[:：]|客户[:：]|LP[:：]|家长[:：]|班主任[:：]|老师[:：])', raw_text)

    current_role = None
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if part in ("员工:", "员工：", "LP:", "LP：", "班主任:", "班主任：", "老师:", "老师："):
            current_role = "lp"
        elif part in ("客户:", "客户：", "家长:", "家长："):
            current_role = "client"
        elif current_role:
            lines.append({"role": current_role, "content": part})

    return lines


def extract_client_messages(conversations: list[dict]) -> list[str]:
    """提取所有客户发言"""
    return [m["content"] for m in conversations if m["role"] == "client"]


def extract_lp_messages(conversations: list[dict]) -> list[str]:
    """提取所有 LP 发言"""
    return [m["content"] for m in conversations if m["role"] == "lp"]


def get_latest_date(dates: list) -> str:
    """获取最近沟通日期"""
    valid = [d for d in dates if d]
    return max(valid) if valid else ""


# ============================================================
# 信号匹配与评分
# ============================================================

def match_signals(client_text: str) -> list[dict]:
    """
    在客户文本中匹配所有信号。
    返回：[{"signal_name":..., "weight":..., "category":..., "matched_keyword":...}, ...]
    """
    matched = []
    text_lower = client_text.lower()

    for sig in ALL_SIGNALS:
        for kw in sig.keywords:
            if kw.lower() in text_lower:
                matched.append({
                    "signal_name": sig.name,
                    "weight": sig.weight,
                    "category": sig.category,
                    "description": sig.description,
                    "matched_keyword": kw,
                })
                break  # 一个信号只计一次

    return matched


def score_student(student_id: str, conversations: list[dict], lp_name: str = "") -> dict:
    """
    方案E：分类上限法。
    在所有对话中匹配信号，去重后按类别计分：
      - 高意向信号: 每类 +1, 上限 8
      - 中意向信号: 每类 +0.5, 上限 3
      - 风险信号: 每类 -1, 下限 -4
    """
    all_signals = []
    all_client_texts = []

    # 逐条分析客户消息
    for conv in conversations:
        if conv["role"] != "client":
            continue
        text = conv["content"]
        all_client_texts.append(text)
        signals = match_signals(text)
        all_signals.extend(signals)

    # 去重信号名，分类
    seen_names = set()
    unique_signals = []
    for s in all_signals:
        if s["signal_name"] not in seen_names:
            seen_names.add(s["signal_name"])
            unique_signals.append(s)

    # 方案E: 分类计数
    high_count = sum(1 for s in unique_signals if s["weight"] > 0 and s["category"] in ("price", "planning", "commit"))
    med_count  = sum(1 for s in unique_signals if s["weight"] > 0 and s["category"] == "neutral")
    risk_count = sum(1 for s in unique_signals if s["weight"] < 0)

    # 方案E 得分
    h_score = min(high_count, 8) * 1.0
    m_score = min(med_count, 6) * 0.5
    r_score = max(risk_count * -1, -4)
    total_score = round(h_score + m_score + r_score, 1)

    # 判定等级
    level = classify_level(total_score, unique_signals)

    # 信号摘要
    positive_signals = [s["signal_name"] for s in unique_signals if s["weight"] > 0]
    risk_signals = [s["signal_name"] for s in unique_signals if s["weight"] < 0]
    low_signals = [s["signal_name"] for s in unique_signals if s["weight"] == 0]

    # 按类别汇总（保留兼容）
    by_category = defaultdict(int)
    for s in unique_signals:
        by_category[s["category"]] += 1

    # 最新一条客户消息
    latest_client_msg = all_client_texts[-1][:200] if all_client_texts else ""

    return {
        "student_id": student_id,
        "lp_name": lp_name,
        "total_messages": len(all_client_texts),
        "total_score": total_score,
        "level": level,
        "level_emoji": {"HIGH": "🔥", "MEDIUM": "🟡", "LOW": "🔵", "RISK": "🔴"}.get(level, "⚪"),
        "positive_signals": positive_signals,
        "risk_signals": risk_signals,
        "low_signals": low_signals,
        "positive_weight": sum(s["weight"] for s in unique_signals if s["weight"] > 0),
        "risk_weight": sum(s["weight"] for s in unique_signals if s["weight"] < 0),
        "latest_client_msg": latest_client_msg,
        "by_category": dict(by_category),
        "all_signals_detail": unique_signals,
    }


def classify_level(total_score: float, signals: list[dict]) -> str:
    """方案E: 根据总分和特殊规则判定意向等级"""
    # 特殊规则1：已发起退款 → 直接风险
    for s in signals:
        if s["signal_name"] == "已发起退款":
            return "RISK"
        if s["signal_name"] == "明确拒绝续费":
            return "RISK"

    # 方案E 阈值
    if total_score >= LEVEL_THRESHOLDS["HIGH"]:   # ≥3
        return "HIGH"
    elif total_score >= LEVEL_THRESHOLDS["MEDIUM"]:  # 1.5~2.9
        return "MEDIUM"
    elif total_score <= LEVEL_THRESHOLDS["RISK"]:   # ≤-1
        return "RISK"
    else:
        return "LOW"  # -0.5 ~ 1.4


# ============================================================
# Excel 读取
# ============================================================

def read_excel(filepath: str) -> list[dict]:
    """读取 Excel，返回学员维度的对话数据"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 找表头行（含"学员ID""LP""材料全对话""沟通时间""过程id"等）
    # 必须同时包含"学员ID"和至少2个其他业务列，避免匹配到过滤参数行
    header_row = None
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        row_vals = [str(c).strip() if c else "" for c in row]
        has_sid = "学员ID" in row_vals
        business_cols = {"LP", "材料全对话", "沟通时间", "小组", "学员姓名", "场景名称", "语义点"}
        overlap = sum(1 for h in row_vals if any(bc in h for bc in business_cols))
        if has_sid and overlap >= 2:
            header_row = row_idx
            headers = row_vals
            break

    if header_row is None:
        # Try fallback: any row with "学员ID" and enough non-empty cells
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            row_vals = [str(c).strip() if c else "" for c in row]
            non_empty = [v for v in row_vals if v]
            if "学员ID" in row_vals and len(non_empty) >= 5:
                header_row = row_idx
                headers = row_vals
                break

    if header_row is None:
        # 按索引读取
        headers = ["学员ID", "区域等级", "LP", "过程id", "沟通时间", "语义点", "材料全对话", "是否执行"]

    # 列索引映射
    col_map = {}
    for i, h in enumerate(headers):
        for key in ["学员ID", "LP", "材料全对话", "沟通时间", "区域等级", "过程id", "语义点", "小组"]:
            if key in str(h):
                col_map[key] = i

    required = ["学员ID", "材料全对话"]
    for r in required:
        if r not in col_map:
            wb.close()
            raise ValueError(f"缺少必要列: {r}。找到的列: {headers}")

    # 按学员聚合对话
    student_data = defaultdict(list)
    lp_map = {}
    date_map = {}
    area_map = {}
    group_map = {}
    student_keys = set()

    def safe_col(row, idx):
        """安全获取列值，处理行长度不一致"""
        return row[idx] if idx < len(row) else ""

    data_start = header_row + 1 if header_row else 1
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        sid = str(safe_col(row, col_map["学员ID"])).strip() if safe_col(row, col_map["学员ID"]) else ""
        if not sid or sid == "学员ID":
            continue

        student_keys.add(sid)
        raw_text = str(safe_col(row, col_map["材料全对话"])) if safe_col(row, col_map["材料全对话"]) else ""
        conversations = parse_conversations(raw_text)

        # 记录 LP 名字
        if "LP" in col_map:
            lp_val = safe_col(row, col_map["LP"])
            if lp_val:
                lp_map[sid] = str(lp_val).strip()

        # 记录沟通时间
        if "沟通时间" in col_map:
            dt_val = safe_col(row, col_map["沟通时间"])
            if dt_val:
                date_map[sid] = str(dt_val).strip()

        # 记录区域
        if "区域等级" in col_map:
            area_val = safe_col(row, col_map["区域等级"])
            if area_val:
                area_map[sid] = str(area_val).strip()

        # 记录小组
        if "小组" in col_map:
            group_val = safe_col(row, col_map["小组"])
            if group_val:
                group_map[sid] = str(group_val).strip()

        # 添加所有对话行
        for conv in conversations:
            student_data[sid].append({
                "id": sid,
                "role": conv["role"],
                "content": conv["content"],
            })

    wb.close()

    # 组装结果
    results = []
    for sid in student_data:
        results.append({
            "student_id": sid,
            "lp_name": lp_map.get(sid, ""),
            "area": area_map.get(sid, ""),
            "group": group_map.get(sid, ""),
            "latest_date": date_map.get(sid, ""),
            "conversations": student_data[sid],
        })

    return results


# ============================================================
# Excel 输出
# ============================================================

def generate_excel(results: list[dict], output_path: str):
    """生成可筛选的学员意向明细 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "续费意向盘"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    level_fills = {
        "HIGH": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # 绿色
        "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # 黄色
        "LOW": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),       # 蓝色
        "RISK": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),      # 红色
    }

    level_fonts = {
        "HIGH": Font(name="微软雅黑", bold=True, color="006100"),
        "MEDIUM": Font(name="微软雅黑", bold=True, color="9C6500"),
        "LOW": Font(name="微软雅黑", bold=True, color="1F4E79"),
        "RISK": Font(name="微软雅黑", bold=True, color="9C0006"),
    }

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    data_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = [
        "序号", "学员ID", "归属LP", "小组", "区域", "意向等级",
        "综合得分", "正向信号数", "风险信号数", "总消息数",
        "正向信号", "风险信号", "最近消息摘要", "跟进行动建议"
    ]
    col_widths = [6, 16, 12, 12, 10, 10, 10, 10, 10, 10, 30, 25, 30, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 按等级排序：RISK → HIGH → MEDIUM → LOW
    level_order = {"RISK": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    results_sorted = sorted(results, key=lambda r: (level_order.get(r["level"], 99), -r["total_score"]))

    # 行动建议映射
    action_map = {
        "HIGH": "优先级P0: 48h内深度沟通推动成交，确认最优方案并引导付款",
        "MEDIUM": "优先级P1: 3天内挖掘顾虑点，结合活动/涨价窗口培育",
        "LOW": "优先级P2: 保持常规触达（每周1次打卡/活动），等待时机",
        "RISK": "优先级P0: 上报LP主管，48h内介入挽留；如确认流失则记录归因",
    }

    for row_idx, r in enumerate(results_sorted, 2):
        values = [
            row_idx - 1,
            r["student_id"],
            r["lp_name"],
            r.get("group", ""),
            r.get("area", ""),
            f'{r["level_emoji"]} {r["level"]}',
            r["total_score"],
            len(r["positive_signals"]),
            len(r["risk_signals"]),
            r.get("total_messages", 0),
            "、".join(r["positive_signals"][:5]) if r["positive_signals"] else "—",
            "、".join(r["risk_signals"][:5]) if r["risk_signals"] else "—",
            r.get("latest_client_msg", "")[:100],
            action_map.get(r["level"], "—"),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = data_align if col_idx >= 10 else center_align

            # 等级列着色
            if col_idx == 5:
                level = r["level"]
                cell.fill = level_fills.get(level, PatternFill())
                cell.font = level_fonts.get(level, Font(name="微软雅黑"))
                cell.alignment = center_align

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results_sorted) + 1}"

    # ---- Sheet 2: 统计汇总 ----
    ws2 = wb.create_sheet("统计汇总")
    level_counts = defaultdict(int)
    for r in results_sorted:
        level_counts[r["level"]] += 1

    total = len(results_sorted)

    summary_headers = ["指标", "数值", "占比"]
    for i, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    summary_data = [
        ("学员总数", total, "100%"),
        ("🔥 高意向", level_counts.get("HIGH", 0), f"{level_counts.get('HIGH', 0)/total*100:.1f}%" if total else "0%"),
        ("🟡 中意向", level_counts.get("MEDIUM", 0), f"{level_counts.get('MEDIUM', 0)/total*100:.1f}%" if total else "0%"),
        ("🔵 观望", level_counts.get("LOW", 0), f"{level_counts.get('LOW', 0)/total*100:.1f}%" if total else "0%"),
        ("🔴 风险", level_counts.get("RISK", 0), f"{level_counts.get('RISK', 0)/total*100:.1f}%" if total else "0%"),
    ]
    for row_idx, (label, val, pct) in enumerate(summary_data, 2):
        ws2.cell(row=row_idx, column=1, value=label).border = thin_border
        ws2.cell(row=row_idx, column=2, value=val).border = thin_border
        ws2.cell(row=row_idx, column=2).alignment = center_align
        ws2.cell(row=row_idx, column=3, value=pct).border = thin_border
        ws2.cell(row=row_idx, column=3).alignment = center_align

    # ---- Sheet 3: LP 维度统计 ----
    ws3 = wb.create_sheet("LP维度统计")
    lp_data = defaultdict(lambda: {"total": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "RISK": 0})
    for r in results_sorted:
        lp = r["lp_name"] or "未知LP"
        lp_data[lp]["total"] += 1
        lp_data[lp][r["level"]] += 1

    lp_headers = ["LP", "学员总数", "🔥高意向", "🟡中意向", "🔵观望", "🔴风险", "高意向率", "风险率"]
    for i, h in enumerate(lp_headers, 1):
        cell = ws3.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for i in range(1, len(lp_headers) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 12
    ws3.column_dimensions["A"].width = 16

    lp_sorted = sorted(lp_data.items(), key=lambda x: x[1]["total"], reverse=True)
    for row_idx, (lp, stats) in enumerate(lp_sorted, 2):
        t = stats["total"]
        high_rate = f"{stats['HIGH']/t*100:.1f}%" if t else "0%"
        risk_rate = f"{stats['RISK']/t*100:.1f}%" if t else "0%"
        for col_idx, val in enumerate([lp, t, stats["HIGH"], stats["MEDIUM"], stats["LOW"], stats["RISK"], high_rate, risk_rate], 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align

    wb.save(output_path)
    print(f"[+] Excel 已保存: {output_path}")


# ============================================================
# Markdown 报告
# ============================================================

def generate_report(results: list[dict], output_path: str, source_file: str):
    """生成管理者 Markdown 报告"""
    now = datetime.now()
    total = len(results)

    level_counts = defaultdict(int)
    for r in results:
        level_counts[r["level"]] += 1

    high = level_counts.get("HIGH", 0)
    medium = level_counts.get("MEDIUM", 0)
    low = level_counts.get("LOW", 0)
    risk = level_counts.get("RISK", 0)

    # 高意向 Top 10
    high_students = sorted(
        [r for r in results if r["level"] == "HIGH"],
        key=lambda x: -x["total_score"]
    )[:10]

    # 风险 Top 10
    risk_students = sorted(
        [r for r in results if r["level"] == "RISK"],
        key=lambda x: x["total_score"]
    )[:10]

    # LP 维度
    lp_data = defaultdict(lambda: {"total": 0, "HIGH": 0, "RISK": 0})
    for r in results:
        lp = r["lp_name"] or "未知LP"
        lp_data[lp]["total"] += 1
        if r["level"] == "HIGH":
            lp_data[lp]["HIGH"] += 1
        if r["level"] == "RISK":
            lp_data[lp]["RISK"] += 1

    # 小组维度
    group_data = defaultdict(lambda: {"total": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "RISK": 0})
    for r in results:
        group = r.get("group") or "未知组"
        group_data[group]["total"] += 1
        lv = r["level"]
        if lv in ("HIGH", "MEDIUM", "LOW", "RISK"):
            group_data[group][lv] += 1

    # 高频风险信号
    risk_signal_counts = defaultdict(int)
    for r in results:
        for s in r.get("risk_signals", []):
            risk_signal_counts[s] += 1
    top_risk_signals = sorted(risk_signal_counts.items(), key=lambda x: -x[1])[:5]

    # 高频正向信号
    pos_signal_counts = defaultdict(int)
    for r in results:
        for s in r.get("positive_signals", []):
            pos_signal_counts[s] += 1
    top_pos_signals = sorted(pos_signal_counts.items(), key=lambda x: -x[1])[:5]

    lines = []

    lines.append(f"# VIPTHINK LP 续费意向盘 — 管理报告")
    lines.append(f"")
    lines.append(f"> 生成时间: {now.strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"> 数据来源: {Path(source_file).name}")
    lines.append(f"> 分析学员: {total} 人")
    lines.append(f"")

    # ---- 一、总览 ----
    lines.append(f"## 一、意向总览")
    lines.append(f"")
    lines.append(f"| 等级 | 人数 | 占比 | 行动要求 |")
    lines.append(f"|------|------|------|---------|")
    lines.append(f"| 🔥 高意向 | {high} | {high/total*100:.1f}% | 48h 内重点跟进成交 |")
    lines.append(f"| 🟡 中意向 | {medium} | {medium/total*100:.1f}% | 3 天内深度沟通培育 |")
    lines.append(f"| 🔵 观望 | {low} | {low/total*100:.1f}% | 保持常规触达节奏 |")
    lines.append(f"| 🔴 风险 | {risk} | {risk/total*100:.1f}% | 升级主管介入挽留 |")
    lines.append(f"")

    # 关键判断
    lines.append(f"### 关键判断")
    lines.append(f"")
    intention_rate = (high + medium) / total * 100 if total else 0
    if intention_rate >= 30:
        lines.append(f"- ✅ 整体续费意向健康（高+中意向率 {intention_rate:.1f}%），重点转化高意向客户")
    elif intention_rate >= 15:
        lines.append(f"- ⚠️ 续费意向中等（高+中意向率 {intention_rate:.1f}%），需加大培育力度")
    else:
        lines.append(f"- 🔴 续费意向偏低（高+中意向率 {intention_rate:.1f}%），需紧急排查流失原因")

    if risk / total > 0.15:
        lines.append(f"- 🔴 风险客户占比 {risk/total*100:.1f}%，超过15%警戒线，需系统性排查")
    lines.append(f"")

    # ---- 二、高意向客户 ----
    lines.append(f"## 二、🔥 高意向客户（共 {high} 人）")
    lines.append(f"")
    if high_students:
        lines.append(f"| 排名 | 学员ID | 归属LP | 得分 | 关键正向信号 |")
        lines.append(f"|------|--------|--------|------|------------|")
        for i, s in enumerate(high_students, 1):
            signals = "、".join(s["positive_signals"][:3])
            lines.append(f"| {i} | {s['student_id']} | {s['lp_name']} | {s['total_score']} | {signals} |")
    else:
        lines.append(f"> 暂无高意向客户")
    lines.append(f"")

    lines.append(f"### 跟进建议")
    lines.append(f"")
    lines.append(f"1. 按得分排序，Top 3 高意向客户应优先电话沟通，而非文字")
    lines.append(f"2. 结合当前活动窗口（如母亲节/520/V9涨价），制造紧迫感")
    lines.append(f"3. 优先推荐方案二（一年课包），价格锚点策略")
    lines.append(f"4. 对询价类客户，直接发价格方案并附「支持全额退款」打消顾虑")
    lines.append(f"")

    # ---- 三、风险客户 ----
    lines.append(f"## 三、🔴 风险客户（共 {risk} 人）")
    lines.append(f"")
    if risk_students:
        lines.append(f"| 排名 | 学员ID | 归属LP | 得分 | 关键风险信号 | 建议动作 |")
        lines.append(f"|------|--------|--------|------|------------|---------|")
        for i, s in enumerate(risk_students, 1):
            risks = "、".join(s["risk_signals"][:3])
            # 根据风险类型给建议
            if "已发起退款" in s["risk_signals"] or "询问退款" in s["risk_signals"]:
                action = "退费专员介入"
            elif "投诉" in str(s["risk_signals"]):
                action = "主管致歉+方案补偿"
            elif "明确拒绝" in str(s["risk_signals"]):
                action = "换LP触达+活动挽回"
            elif "课程不满" in str(s["risk_signals"]):
                action = "安排试高级别课程"
            else:
                action = "电话沟通确认意向"
            lines.append(f"| {i} | {s['student_id']} | {s['lp_name']} | {s['total_score']} | {risks} | {action} |")
    else:
        lines.append(f"> 暂无风险客户")
    lines.append(f"")

    # ---- 四、风险信号分布 ----
    lines.append(f"## 四、风险信号分布 TOP 5")
    lines.append(f"")
    lines.append(f"| 风险信号 | 出现人次 |")
    lines.append(f"|---------|---------|")
    for sig, cnt in top_risk_signals:
        lines.append(f"| {sig} | {cnt} |")
    lines.append(f"")

    # ---- 五、LP 维度分析 ----
    lines.append(f"## 五、LP 维度分析")
    lines.append(f"")
    lines.append(f"| LP | 学员数 | 高意向数 | 高意向率 | 风险数 | 风险率 |")
    lines.append(f"|------|--------|---------|---------|--------|--------|")
    for lp, stats in sorted(lp_data.items(), key=lambda x: x[1]["total"], reverse=True):
        t = stats["total"]
        h_rate = f"{stats['HIGH']/t*100:.1f}%" if t else "0%"
        r_rate = f"{stats['RISK']/t*100:.1f}%" if t else "0%"
        lines.append(f"| {lp} | {t} | {stats['HIGH']} | {h_rate} | {stats['RISK']} | {r_rate} |")
    lines.append(f"")

    # ---- 六、小组维度分析 ----
    lines.append(f"## 六、小组维度分析")
    lines.append(f"")
    lines.append(f"| 小组 | 学员数 | 高意向 | 高意向率 | 中意向 | 观望 | 风险 | 风险率 |")
    lines.append(f"|------|--------|--------|---------|--------|------|------|--------|")
    for group, stats in sorted(group_data.items(), key=lambda x: x[1]["total"], reverse=True):
        t = stats["total"]
        h_rate = f"{stats['HIGH']/t*100:.1f}%" if t else "0%"
        r_rate = f"{stats['RISK']/t*100:.1f}%" if t else "0%"
        lines.append(f"| {group} | {t} | {stats['HIGH']} | {h_rate} | {stats['MEDIUM']} | {stats['LOW']} | {stats['RISK']} | {r_rate} |")
    lines.append(f"")

    # ---- 七、整体建议 ----
    lines.append(f"## 七、管理者行动清单")
    lines.append(f"")
    lines.append(f"### 本月重点工作")
    lines.append(f"")
    lines.append(f"- [ ] 高意向客户 48h 内分配专人跟进（共 {high} 人）")
    lines.append(f"- [ ] 风险客户由部门主管逐一审核（共 {risk} 人），确认流失原因")
    lines.append(f"- [ ] 「{top_risk_signals[0][0] if top_risk_signals else '无明显风险信号'}」是最突出的风险信号，需团队周会讨论应对方案")
    lines.append(f"- [ ] 「{top_pos_signals[0][0] if top_pos_signals else '无明显正向信号'}」是最常见的正向信号，可作为续费话术切入点")
    lines.append(f"")
    lines.append(f"### 下周复盘项")
    lines.append(f"")
    lines.append(f"- [ ] 高意向客户中有多少完成了续费？（转化率）")
    lines.append(f"- [ ] 风险客户中有多少成功挽回？（挽回率）")
    lines.append(f"- [ ] 本次意向盘 vs 实际续费结果的偏差分析")
    lines.append(f"")

    lines.append(f"---")
    lines.append(f"> 本报告由 vipthink-renewal-intention skill 自动生成")

    content = "\n".join(lines)
    Path(output_path).write_text(content, encoding="utf-8")
    print(f"[+] 报告已保存: {output_path}")


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="VIPTHINK LP 续费意向盘分析")
    parser.add_argument("input", help="输入的 Excel 文件路径")
    parser.add_argument("--output-dir", "-o", default=".", help="输出目录 (默认: 当前目录)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[!] 文件不存在: {input_path}")
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    today = date.today().strftime("%Y-%m-%d")

    print(f"[*] 读取数据: {input_path}")
    students = read_excel(str(input_path))
    print(f"[*] 学员数: {len(students)}")

    print(f"[*] 正在分析续费意向...")
    results = []
    for i, stu in enumerate(students):
        if (i + 1) % 500 == 0:
            print(f"  进度: {i+1}/{len(students)}")
        analysis = score_student(
            stu["student_id"],
            stu["conversations"],
            lp_name=stu["lp_name"]
        )
        analysis["lp_name"] = stu["lp_name"]
        analysis["area"] = stu["area"]
        analysis["latest_date"] = stu["latest_date"]
        results.append(analysis)

    # 按等级统计
    level_counts = defaultdict(int)
    for r in results:
        level_counts[r["level"]] += 1
    print(f"[*] 分析完成:")
    print(f"    🔥 高意向: {level_counts.get('HIGH', 0)}")
    print(f"    🟡 中意向: {level_counts.get('MEDIUM', 0)}")
    print(f"    🔵 观望: {level_counts.get('LOW', 0)}")
    print(f"    🔴 风险: {level_counts.get('RISK', 0)}")

    # 生成 Excel
    excel_path = output_dir / f"renewal_intention_{today}.xlsx"
    print(f"[*] 生成 Excel: {excel_path}")
    generate_excel(results, str(excel_path))

    # 生成报告
    report_path = output_dir / f"renewal_report_{today}.md"
    print(f"[*] 生成报告: {report_path}")
    generate_report(results, str(report_path), input_path.name)

    print(f"\n[+] 全部完成！")
    print(f"    📊 Excel: {excel_path}")
    print(f"    📝 报告: {report_path}")


if __name__ == "__main__":
    main()
