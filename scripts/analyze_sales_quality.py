# -*- coding: utf-8 -*-
"""
VIPTHINK LP 销售过程质量分析脚本
=================================
分析 LP-家长对话中的销售过程问题，生成逐学员诊断+建议+统计。

用法:
    python analyze_sales_quality.py <输入Excel> [--output-dir <目录>]
"""
import argparse, os, re, sys
from collections import defaultdict, Counter
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 信号分类（同 analyze_renewal.py）
# ============================================================
HIGH_SIGNALS = [
    ('主动询价', +3, ['多少钱', '价格多少', '怎么收费', '费用', '学费多少', '收费方案', '优惠方案', '收费方案']),
    ('比价/方案比较', +3, ['比较', '对比', '哪个划算', '方案一', '方案二']),
    ('价格敏感/算账', +3, ['每堂', '每节课', '一堂多少', '单价', '折合', '平均']),
    ('升舱/升级咨询', +3, ['升舱', '升级', '下个阶段', '下个级别', '后面学什么']),
    ('讨价还价', +3, ['能不能便宜', '再便宜', '最低多少', '优惠', '折扣']),
    ('主动要方案', +3, ['发我', '给我看看', '方案', '报价']),
    ('正向续费表态', +3, ['续费', '继续学', '接着上', '继续上', '接着学', '再报', '继续报']),
    ('进度关心', +2, ['进度', '学了多久', '学到哪', '还剩多少', '课时']),
    ('付款方式咨询', +2, ['付款', '支付', '定金', '尾款', '分期']),
    ('长期规划', +2, ['长期', '规划', '以后', '将来', '计划']),
    ('课时冻结咨询', +2, ['冻结', '暂停', '请假', '停课']),
    ('价格锁定意愿', +2, ['锁定', '保留', '预留', '名额']),
]
MEDIUM_SIGNALS = [
    ('正面回应', +1, ['好的', '可以', '行', '没问题', '明白', '了解', '嗯嗯', '好呀', 'OK', 'ok']),
    ('意识到涨价', +1, ['涨价', '调价', '贵了', '涨了', '加价']),
    ('认可课程价值', +1, ['进步', '效果', '喜欢', '感兴趣', '有帮助', '不错', '很好', '棒']),
    ('需要商量', +1, ['商量', '问下', '爸爸', '妈妈', '老公', '老婆', '家人']),
    ('表示再考虑', +1, ['考虑', '想想', '再看看', '先看看']),
    ('时间不合适但有意向', +1, ['时间', '暑假', '假期', '忙', '安排']),
]
RISK_SIGNALS = [
    ('明确拒绝续费', -4, ['不续', '不续费', '不学了', '不报', '不买', '不要了', '不考虑了', '不需要']),
    ('课程不满', -3, ['没效果', '没进步', '不好', '不满意', '不喜欢', '无聊', '太难', '太简单']),
    ('有替代方案', -3, ['其他机构', '别的', '换', '线下', '学校有', '自学']),
    ('老师不满', -3, ['老师不好', '换老师', '不喜欢老师']),
    ('系统/服务不满', -3, ['卡', '系统', 'APP', '闪退', '服务差']),
    ('询问退款', -4, ['退款', '退费', '怎么退', '能退吗']),
    ('已发起退款', -4, ['已经退', '退了', '申请退']),
]
LEVEL_THRESHOLDS = {'HIGH': 5, 'MEDIUM': 2, 'LOW': -1, 'RISK': -2}

# ============================================================
# 模板/问题模式定义
# ============================================================
TEMPLATE_PATTERNS = [
    '數理思維進步規劃指南嚟啦', 'VIPTHINK學數學就上VIPTHINK',
    '本週打卡任務', '幫小朋友領取2000豌豆幣',
    '520 限時預售', '數理思維進步規劃指南', '掃碼即刻免費領取',
    '誠意推介大家掃碼免費試聽', '该消息类型暂不能展示',
    '520抽奖', '瓜分 52000', '鎖定升階學位及抽獎活動資格',
    '打卡截圖要上傳或者需要調補課', '老師仲差 2 個推薦名額',
    '邀請2位3-11歲小朋友', '即送6節直播課', '空氣炸鍋', '100節物理啟蒙',
    '52000 豌豆幣', '520限時禮遇', '母親節', '母亲节', '五一', '勞動節',
    '升階學位正在緊張補錄', '原班學位名額非常有限',
]
DAKA_PATTERNS = ['打卡', '打卡任務', '2000豌豆幣', '豌豆幣', '打卡截圖']
REFERRAL_PATTERNS = ['轉介紹', '推荐', '推薦', '介紹朋友', '邀請朋友', '朋友來',
                     '試聽獎勵', '推薦獎勵', '赠课', '贈課', '空氣炸鍋', '物理啟蒙',
                     '邀請有禮', '推薦有禮', '老帶新', '拉新', '朋友試聽', '推荐好友']
REFUND_PATTERNS = ['退費靈活', '隨時退', '無條件退', '全額退款', '不滿意退',
                   '退費保障', '退款保障', '支持退款', '可退', '隨上隨退',
                   '買貴包退', '無違約金', '無手續費']
PRICE_ANCHOR_PATTERNS = ['每堂', '每課', '一堂', '一節', '單價', '原價', '現在',
                         '漲價', '調價', '之後', '以後', '升價', '省', '慳',
                         '元/課', '蚊一堂', '立減', '便宜過']
EMOTIONAL_PATTERNS = ['考核', '評選', '優秀班主任', '幫我', '幫手', '差少少就達成',
                      '指標', '放假', '返鄉下', '拜托', '麻煩幫我']
PAIN_POINT_PATTERNS = ['哪方面', '哪裡', '困難', '問題', '弱項', '不足', '需要加強',
                       '需要提升', '進步點', '帶提升', '粗心', '不專心', '走神']
PERSONAL_PATTERNS = ['小朋友', '寶貝', '宝贝', '進步', '进步', '課堂', '课堂', '老師', '老师反馈']
RENEWAL_PATTERNS = ['續費', '续费', '升階', '升阶', '課包', '课包', '報名', '报名']

# ============================================================
# 对话解析
# ============================================================
def parse_dialogue(raw_text):
    if not raw_text or not isinstance(raw_text, str):
        return []
    parts = re.split(r'(老师[:：]|班主任[:：]|员工[:：]|LP[:：]|家长[:：]|客户[:：])', raw_text.strip())
    lines, current_role = [], None
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if part in ('老师:', '老师：', '班主任:', '班主任：', '员工:', '员工：', 'LP:', 'LP：'):
            current_role = 'lp'
        elif part in ('家长:', '家长：', '客户:', '客户：'):
            current_role = 'parent'
        elif current_role:
            lines.append({'role': current_role, 'content': part})
    return lines

# ============================================================
# 续费意向评分
# ============================================================
def score_intention(parent_text, all_text):
    score, pos_signals, risk_signals = 0, [], []
    has_refund_request, has_explicit_reject = False, False

    for sname, weight, keywords in HIGH_SIGNALS + MEDIUM_SIGNALS:
        if keywords and any(kw in parent_text for kw in keywords):
            pos_signals.append(sname)
            score += weight
    for sname, weight, keywords in RISK_SIGNALS:
        if keywords and any(kw in parent_text for kw in keywords):
            risk_signals.append(sname)
            score += weight
            if sname == '明确拒绝续费':
                has_explicit_reject = True
            if sname in ('询问退款', '已发起退款'):
                has_refund_request = True

    if has_refund_request or has_explicit_reject:
        level = 'RISK'
    elif score >= LEVEL_THRESHOLDS['HIGH']:
        level = 'HIGH'
    elif score >= LEVEL_THRESHOLDS['MEDIUM']:
        level = 'MEDIUM'
    elif score >= LEVEL_THRESHOLDS['LOW']:
        level = 'LOW'
    else:
        level = 'RISK'

    return {'score': score, 'level': level, 'pos_signals': pos_signals, 'risk_signals': risk_signals}

# ============================================================
# LP 销售问题检测
# ============================================================
def detect_problems(parsed, lp_msgs, parent_msgs):
    problems, details = [], {}
    if not parsed:
        problems.append('无对话内容')
        return problems, details

    lp_text = ' '.join(m['content'] for m in lp_msgs)
    parent_text = ' '.join(m['content'] for m in parent_msgs)
    parent_cnt = len(parent_msgs)
    lp_cnt = len(lp_msgs)
    details.update({'parent_reply_count': parent_cnt, 'lp_msg_count': lp_cnt})

    # 模板轰炸
    tc = sum(1 for pat in TEMPLATE_PATTERNS if pat in lp_text)
    details['template_count'] = tc
    if tc >= 5: problems.append('模板轰炸(>=5条)')
    elif tc >= 3: problems.append('模板偏多(3-4条)')

    # 打卡干扰
    if any(pat in lp_text for pat in DAKA_PATTERNS):
        problems.append('打卡任务干扰续费对话')

    # 单向广播
    if parent_cnt == 0:
        problems.append('家长零回复(单向广播)')
    elif parent_cnt <= 2 and lp_cnt >= 5:
        problems.append('家长几乎无互动')

    # 过早放弃
    reject_kw = ['不考慮', '不考虑', '不續', '不续', '不報', '不报', '不買', '不买', '暫時不', '暂时不']
    if any(kw in parent_text for kw in reject_kw):
        found_reject = False
        lp_after = []
        for m in parsed:
            if m['role'] == 'parent' and any(kw in m['content'] for kw in reject_kw):
                found_reject = True
            elif found_reject and m['role'] == 'lp':
                lp_after.append(m['content'])
        if not any(any(kw in m for kw in ['價格', '价格', '方案', '优惠', '試試', '试试', '保障', '放心'])
                   for m in lp_after[:3]):
            problems.append('家长拒绝后未做有效挽救')

    # 转介绍喧宾夺主
    ref_cnt = sum(1 for pat in REFERRAL_PATTERNS if pat in lp_text)
    ren_cnt = sum(1 for pat in RENEWAL_PATTERNS if pat in lp_text)
    details['referral_count'] = ref_cnt
    details['renewal_mention_count'] = ren_cnt
    if ref_cnt > ren_cnt:
        problems.append('转介绍推荐喧宾夺主(>续费提及)')

    # 缺退费保障
    if not any(pat in lp_text for pat in REFUND_PATTERNS) and lp_cnt >= 3:
        problems.append('未使用退费保障打消顾虑')

    # 缺价格锚点
    if not any(pat in lp_text for pat in PRICE_ANCHOR_PATTERNS) and lp_cnt >= 3:
        problems.append('未做价格对比/锚点')

    # 情绪绑架
    if any(pat in lp_text for pat in EMOTIONAL_PATTERNS):
        problems.append('使用情绪施压话术(考核/指标)')

    # 缺痛点挖掘
    if not any(pat in lp_text for pat in PAIN_POINT_PATTERNS) and lp_cnt >= 3:
        problems.append('未做学习痛点/进步点挖掘')

    # 缺个性化
    if not any(pat in lp_text for pat in PERSONAL_PATTERNS) and lp_cnt >= 3:
        problems.append('缺乏个性化沟通(未提学员具体情况)')

    return problems, details

# ============================================================
# 建议生成
# ============================================================
def generate_advice(problems, intention):
    parts = []
    level = intention['level']
    if level == 'HIGH':
        parts.append('【高意向-48h内电话成交】优先电话沟通，直接发方案+退费保障，制造活动紧迫感')
    elif level == 'MEDIUM':
        parts.append('【中意向-3天内深度培育】挖掘学习痛点，展示下阶段内容价值，给价格对比方案')
    elif level == 'LOW':
        parts.append('【观望-保持触达】暂停推销，先发学习报告建立信任，2周后再提')
    elif level == 'RISK':
        parts.append('【风险-升级处理】停止模板消息，主管介入了解原因，针对性解决')

    advice_map = {
        '模板轰炸(>=5条)': '减少群发模板，改为个性化开头"XX妈妈，宝贝最近在XXX方面进步很大..."',
        '模板偏多(3-4条)': '控制模板使用，每条模板前加一句个性化内容',
        '打卡任务干扰续费对话': '续费周期内暂停打卡提醒，避免信息过载',
        '家长零回复(单向广播)': '改为电话触达，文字触达已失效',
        '家长几乎无互动': '减少消息频率，改用提问式开场"宝贝最近上课感觉怎么样？"',
        '家长拒绝后未做有效挽救': '拒绝后追问原因+给替代方案(小课包/试学期/退费保障)',
        '转介绍推荐喧宾夺主(>续费提及)': '续费未成交前暂停转介绍推送，避免反感',
        '未使用退费保障打消顾虑': '补充退费保障"支持随上随退，上一节算一节，无违约金"',
        '未做价格对比/锚点': '补充价格锚点"现在XX元/节，升舱后至少XX元/节，早规划省XXXX元"',
        '使用情绪施压话术(考核/指标)': '停用考核/指标类话术，改用利他型"给朋友送免费试听福利"',
        '未做学习痛点/进步点挖掘': '开头先汇报孩子进步+待提升点，建立专业信任后再提续费',
        '缺乏个性化沟通(未提学员具体情况)': '提及学员具体数据：出勤率、正确率、进步幅度、老师评语',
    }
    for p in problems:
        if p in advice_map:
            parts.append(advice_map[p])
    if '无对话内容' in problems:
        parts = ['【无对话记录】需确认是否有线下沟通，如无则需首次电话触达：汇报学习情况 > 预告升阶 > 发方案']

    return ' | '.join(parts)

# ============================================================
# 主流程
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='VIPTHINK LP 销售过程质量分析')
    parser.add_argument('input', help='输入 Excel 文件路径')
    parser.add_argument('--output-dir', '-o', default=None, help='输出目录 (默认与输入同目录)')
    args = parser.parse_args()

    out_dir = args.output_dir or os.path.dirname(args.input) or '.'
    os.makedirs(out_dir, exist_ok=True)
    today = datetime.now().strftime('%Y-%m-%d')

    print('读取: {}'.format(args.input))
    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active

    # 聚合学员
    students = defaultdict(lambda: {'sid': '', 'name': '', 'lp': '', 'group': '', 'dialogues': []})
    for row_idx in range(2, ws.max_row + 1):
        sid = str(ws.cell(row=row_idx, column=1).value or '').strip()
        if not sid:
            continue
        s = students[sid]
        s['sid'] = sid
        s['name'] = str(ws.cell(row=row_idx, column=2).value or '').strip()
        s['lp'] = str(ws.cell(row=row_idx, column=3).value or '').strip()
        s['group'] = str(ws.cell(row=row_idx, column=4).value or '').strip()
        s['dialogues'].append(str(ws.cell(row=row_idx, column=21).value or ''))
    wb.close()

    print('去重学员: {} 人'.format(len(students)))

    results = []
    lp_problem_stats = defaultdict(lambda: defaultdict(int))
    lp_total = defaultdict(int)

    for sid, s in students.items():
        lp = s['lp']
        lp_total[lp] += 1

        all_text = ' '.join(s['dialogues'])
        parsed = []
        for d in s['dialogues']:
            parsed.extend(parse_dialogue(d))

        lp_msgs = [m for m in parsed if m['role'] == 'lp']
        parent_msgs = [m for m in parsed if m['role'] == 'parent']
        parent_text = ' '.join(m['content'] for m in parent_msgs)

        intention = score_intention(parent_text, all_text)
        problems, details = detect_problems(parsed, lp_msgs, parent_msgs)
        for p in problems:
            lp_problem_stats[lp][p] += 1

        advice = generate_advice(problems, intention)

        results.append({
            'sid': sid, 'name': s['name'], 'lp': lp, 'group': s['group'],
            'level': intention['level'], 'score': intention['score'],
            'pos': '、'.join(intention['pos_signals']) or '-',
            'risk': '、'.join(intention['risk_signals']) or '-',
            'problems': '；'.join(problems) if problems else '无明显问题',
            'problem_count': len(problems),
            'lp_cnt': details.get('lp_msg_count', 0),
            'parent_cnt': details.get('parent_reply_count', 0),
            'template_cnt': details.get('template_count', 0),
            'advice': advice,
        })

    # ==== 写 Excel ====
    out_wb = openpyxl.Workbook()
    hfill = openpyxl.styles.PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hfont = openpyxl.styles.Font(bold=True, color='FFFFFF', size=11)
    lcolors = {
        'HIGH': openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'MEDIUM': openpyxl.styles.PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'LOW': openpyxl.styles.PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        'RISK': openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    # Sheet 1: 学员综合分析
    ws1 = out_wb.active
    ws1.title = '学员综合分析'
    h1 = ['序号', '学员ID', '学员姓名', 'LP', '小组', '续费意向', '得分', '正向信号', '风险信号',
          'LP问题点', '问题数', 'LP发消息', '家长回复', '模板数', '建议谈单思路']
    for ci, h in enumerate(h1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font, c.fill = hfont, hfill
        c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    lorder = {'RISK': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
    results.sort(key=lambda x: (lorder.get(x['level'], 9), -x['score']))
    for i, r in enumerate(results, 1):
        row = [i, r['sid'], r['name'], r['lp'], r['group'], r['level'], r['score'],
               r['pos'], r['risk'], r['problems'], r['problem_count'], r['lp_cnt'],
               r['parent_cnt'], r['template_cnt'], r['advice']]
        for ci, v in enumerate(row, 1):
            cell = ws1.cell(row=i+1, column=ci, value=v)
        if r['level'] in lcolors:
            ws1.cell(row=i+1, column=6).fill = lcolors[r['level']]

    widths1 = [5, 12, 12, 8, 8, 10, 6, 25, 20, 45, 6, 10, 10, 8, 55]
    for ci, w in enumerate(widths1, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws1.freeze_panes, ws1.auto_filter.ref = 'A2', 'A1:O{}'.format(len(results)+1)

    # Sheet 2: 意向统计
    ws2 = out_wb.create_sheet('意向统计')
    lc = Counter(r['level'] for r in results)
    total = len(results)
    for ri, (label, key) in enumerate([('等级', ''), ('高意向', 'HIGH'), ('中意向', 'MEDIUM'),
                                        ('观望', 'LOW'), ('风险', 'RISK'), ('合计', '')], 1):
        if ri == 1:
            for ci, h in enumerate(['等级', '人数', '占比'], 1):
                c = ws2.cell(row=1, column=ci, value=h); c.font, c.fill = hfont, hfill
        elif key:
            ws2.cell(row=ri, column=1, value=label)
            ws2.cell(row=ri, column=2, value=lc.get(key, 0))
            ws2.cell(row=ri, column=3, value='{:.1f}%'.format(lc.get(key,0)/total*100))
            if key in lcolors:
                ws2.cell(row=ri, column=1).fill = lcolors[key]
        else:
            ws2.cell(row=ri, column=1, value='合计').font = openpyxl.styles.Font(bold=True)
            ws2.cell(row=ri, column=2, value=total).font = openpyxl.styles.Font(bold=True)
            ws2.cell(row=ri, column=3, value='100%').font = openpyxl.styles.Font(bold=True)

    # Sheet 3: LP 问题统计
    ws3 = out_wb.create_sheet('LP问题统计')
    all_probs = sorted(set(p for probs in lp_problem_stats.values() for p in probs))
    ws3.cell(row=1, column=1, value='LP').font = hfont; ws3.cell(row=1, column=1).fill = hfill
    for ci, prob in enumerate(all_probs, 2):
        c = ws3.cell(row=1, column=ci, value=prob); c.font, c.fill = hfont, hfill
    for ri, lp in enumerate(sorted(lp_total.keys()), 2):
        ws3.cell(row=ri, column=1, value=lp)
        for ci, prob in enumerate(all_probs, 2):
            ws3.cell(row=ri, column=ci, value=lp_problem_stats[lp].get(prob, 0))
    ws3.freeze_panes = 'B2'

    # Sheet 4: 问题类型汇总
    ws4 = out_wb.create_sheet('问题类型汇总')
    pc = Counter()
    plp = defaultdict(set)
    for lp, probs in lp_problem_stats.items():
        for p, cnt in probs.items():
            pc[p] += cnt; plp[p].add(lp)
    for ci, h in enumerate(['问题类型', '出现人次', '占比', '涉及LP数', 'LP占比'], 1):
        c = ws4.cell(row=1, column=ci, value=h); c.font, c.fill = hfont, hfill
    for ri, (prob, cnt) in enumerate(sorted(pc.items(), key=lambda x: -x[1]), 2):
        ws4.cell(row=ri, column=1, value=prob)
        ws4.cell(row=ri, column=2, value=cnt)
        ws4.cell(row=ri, column=3, value='{:.1f}%'.format(cnt/total*100))
        ws4.cell(row=ri, column=4, value=len(plp[prob]))
        ws4.cell(row=ri, column=5, value='{:.1f}%'.format(len(plp[prob])/len(lp_total)*100))
    for ci, w in enumerate([35, 10, 10, 10, 10], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out_path = os.path.join(out_dir, 'renewal_quality_{}.xlsx'.format(today))
    out_wb.save(out_path)

    print('\n输出: {}'.format(out_path))
    print('学员: {} | HIGH={} MEDIUM={} LOW={} RISK={}'.format(
        total, lc.get('HIGH',0), lc.get('MEDIUM',0), lc.get('LOW',0), lc.get('RISK',0)))
    print('\n问题 Top 5:')
    for prob, cnt in sorted(pc.items(), key=lambda x: -x[1])[:5]:
        print('  {}: {}人次 ({} LP)'.format(prob, cnt, len(plp[prob])))

if __name__ == '__main__':
    main()
