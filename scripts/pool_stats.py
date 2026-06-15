#!/usr/bin/env python3
"""续费池子拆分统计: 按池子→学员数/高意向/签单/风险

用法:
  python pool_stats.py <原始BI.xlsx> [--output-dir .]
"""
import argparse, os, sys
from collections import defaultdict
from datetime import date
import openpyxl

POOL_NAMES = {
    "早鸟": "早鸟",
    "当月结课未续": "当月结课未续",
    "次月结课未续": "次月结课未续",
    "次次月结课未续": "次次月结课未续",
    "升舱未续": "升舱未续",
    "活跃低课时未续": "活跃低课时未续",
}


def main():
    parser = argparse.ArgumentParser(description="池子拆分统计")
    parser.add_argument("raw_excel")
    parser.add_argument("--output-dir", default=".")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.raw_excel, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, min(12, ws.max_row + 1)):
        row_vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
        if "学员ID" in row_vals and "池子" in row_vals:
            header_row = r
            break

    if not header_row:
        print("ERROR: 找不到表头", file=sys.stderr)
        sys.exit(1)

    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
    sid_col = next((i for i, h in enumerate(headers) if h == "学员ID"), 0)
    pool_col = next((i for i, h in enumerate(headers) if h == "池子"), 15)
    signed_col = next((i for i, h in enumerate(headers) if "是否续费" in h), 8)

    # Aggregate: {sid: {pool, signed}}
    students = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sid = str(row[sid_col]).strip().rstrip(".0") if row[sid_col] else ""
        if not sid or sid == "None":
            continue
        if sid not in students:
            pool = str(row[pool_col]).strip() if row[pool_col] else "未知"
            signed = 1 if row[signed_col] == 1.0 or str(row[signed_col]).strip() == "1" else 0
            students[sid] = {"pool": pool, "signed": signed}
    wb.close()

    # Pool stats
    pool_data = defaultdict(lambda: {"学员数": 0, "已签": 0, "高意向": 0, "风险": 0})
    
    # Try to merge with intent data for HIGH/RISK
    intent_path = os.path.join(args.output_dir, f"renewal_intention_{date.today()}.xlsx")
    intent_map = {}
    if os.path.exists(intent_path):
        iwb = openpyxl.load_workbook(intent_path, data_only=True)
        iws = iwb.active
        ih = [str(c.value).strip() if c.value else "" for c in iws[1]]
        isid = next((i for i, h in enumerate(ih) if "学员ID" in h), 1)
        ilv = next((i for i, h in enumerate(ih) if "意向等级" in h), 5)
        for row in iws.iter_rows(min_row=2, values_only=True):
            sid = str(row[isid]).strip().rstrip(".0") if row[isid] else ""
            lv = str(row[ilv]).strip() if row[ilv] else ""
            if sid:
                intent_map[sid] = lv
        iwb.close()

    for sid, info in students.items():
        pool = info["pool"]
        pool_data[pool]["学员数"] += 1
        if info["signed"]:
            pool_data[pool]["已签"] += 1
        lv = intent_map.get(sid, "")
        if "HIGH" in lv or "高" in lv:
            pool_data[pool]["高意向"] += 1
        elif "RISK" in lv or "风险" in lv:
            pool_data[pool]["风险"] += 1

    # Write output
    out_wb = openpyxl.Workbook()
    ws_out = out_wb.active
    ws_out.title = "池子统计"
    hfill = openpyxl.styles.PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfont = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)

    hdrs = ["池子", "学员数", "已签", "签单率", "高意向", "高风险", "未跟进占比"]
    for ci, h in enumerate(hdrs, 1):
        c = ws_out.cell(row=1, column=ci, value=h)
        c.font, c.fill = hfont, hfill

    ordered_pools = list(POOL_NAMES.keys()) + ["未知"]
    for ri, pool in enumerate(ordered_pools, 2):
        if pool not in pool_data:
            continue
        d = pool_data[pool]
        total = d["学员数"]
        sign_rate = f"{d['已签']/total*100:.1f}%" if total else "0%"
        risk_rate = f"{d['风险']/total*100:.1f}%" if total else "0%"
        ws_out.cell(row=ri, column=1, value=pool)
        ws_out.cell(row=ri, column=2, value=total)
        ws_out.cell(row=ri, column=3, value=d["已签"])
        ws_out.cell(row=ri, column=4, value=sign_rate)
        ws_out.cell(row=ri, column=5, value=d["高意向"])
        ws_out.cell(row=ri, column=6, value=f"{d['风险']}({risk_rate})")

    for ci, w in enumerate([18, 10, 8, 10, 10, 14, 14], 1):
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out_path = os.path.join(args.output_dir, f"pool_stats_{date.today()}.xlsx")
    out_wb.save(out_path)

    print(f"池子统计: {out_path}")
    total_students = sum(d["学员数"] for d in pool_data.values())
    total_signed = sum(d["已签"] for d in pool_data.values())
    print(f"总计: {total_students}人, 已签{total_signed}人 ({total_signed/max(total_students,1)*100:.1f}%)")
    for pool in ordered_pools:
        if pool in pool_data:
            d = pool_data[pool]
            print(f"  {pool}: {d['学员数']}人 签{d['已签']}({d['已签']/max(d['学员数'],1)*100:.1f}%) HIGH={d['高意向']} RISK={d['风险']}")

    # Return dict for downstream use
    return {
        "pools": {p: dict(pool_data[p]) for p in pool_data},
        "total": total_students,
        "signed": total_signed,
    }


if __name__ == "__main__":
    main()
