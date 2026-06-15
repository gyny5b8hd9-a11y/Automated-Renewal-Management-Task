#!/usr/bin/env python3
"""对续费意向分析结果按小组维度做二次统计"""
import argparse
from pathlib import Path
from collections import defaultdict
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def main():
    parser = argparse.ArgumentParser(description="按小组汇总续费意向分析结果")
    parser.add_argument("raw_data", help="原始导出的Excel路径（含小组列）")
    parser.add_argument("intention_file", help="续费意向分析Excel路径")
    parser.add_argument("--output-dir", help="输出目录", default=None)
    args = parser.parse_args()

    raw_path = Path(args.raw_data)
    intention_path = Path(args.intention_file)
    out_dir = Path(args.output_dir) if args.output_dir else intention_path.parent
    today = date.today()

    # 1. 读取原始数据获取小组映射（学员ID → 小组）
    print(f"[*] 读取原始数据: {raw_path}")
    df_raw = pd.read_excel(raw_path, dtype=str)
    # 找到表头行
    header_idx = None
    for i in range(min(15, len(df_raw))):
        row_vals = [str(v).strip() for v in df_raw.iloc[i].values]
        if "学员ID" in row_vals and sum(1 for h in row_vals if any(k in str(h) for k in ["小组", "LP", "材料全对话"])) >= 2:
            header_idx = i
            break
    if header_idx is None:
        print("[!] 找不到表头行，尝试第4行")
        header_idx = 4

    df_raw.columns = [str(c).strip() for c in df_raw.iloc[header_idx].values]
    df_raw = df_raw.iloc[header_idx + 1:].reset_index(drop=True)

    sid_col = next((c for c in df_raw.columns if "学员ID" in str(c)), "学员ID")
    group_col = next((c for c in df_raw.columns if "小组" in str(c)), "小组")

    # Build SID → group map (keep last occurrence)
    sid_to_group = {}
    for _, row in df_raw.iterrows():
        sid = str(row[sid_col]).strip()
        gp = str(row[group_col]).strip()
        if sid and sid != "nan":
            sid_to_group[sid] = gp if gp and gp != "nan" else "未知组"

    print(f"[*] 小组映射: {len(sid_to_group)} 学员ID, {len(set(sid_to_group.values()))} 个不同小组")

    # 2. 读取意向分析结果
    print(f"[*] 读取意向分析: {intention_path}")
    df_intent = pd.read_excel(intention_path, sheet_name="续费意向盘", dtype=str)
    
    # 添加小组列
    def find_group(sid_val):
        sid = str(sid_val).strip().rstrip(".0")  # Remove float suffix
        # Try with and without .0
        for try_sid in [sid, sid + ".0"]:
            if try_sid in sid_to_group:
                return sid_to_group[try_sid]
        return "未知组"

    df_intent["小组"] = df_intent["学员ID"].apply(find_group)

    # 3. 按小组统计
    print(f"[*] 按小组汇总统计...")
    group_stats = defaultdict(lambda: {"学员数": 0, "高意向": 0, "中意向": 0, "观望": 0, "风险": 0})

    level_map = {}
    if "意向等级" in df_intent.columns:
        for _, row in df_intent.iterrows():
            gp = str(row.get("小组", "")).strip()
            lv = str(row.get("意向等级", "")).strip()
            group_stats[gp]["学员数"] += 1
            if "HIGH" in lv or "高" in lv:
                group_stats[gp]["高意向"] += 1
            elif "MED" in lv or "中" in lv:
                group_stats[gp]["中意向"] += 1
            elif "RISK" in lv or "风险" in lv:
                group_stats[gp]["风险"] += 1
            else:
                group_stats[gp]["观望"] += 1

    # 4. 输出Excel
    out_excel = out_dir / f"group_stats_{today}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "小组统计"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    data_align = Alignment(horizontal="center", vertical="center")

    headers = ["小组", "学员数", "高意向", "高意向率", "中意向", "观望", "风险", "风险率"]
    col_widths = [16, 10, 10, 12, 10, 10, 10, 10]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    sorted_groups = sorted(group_stats.items(), key=lambda x: x[1]["学员数"], reverse=True)
    for row_idx, (gp, stats) in enumerate(sorted_groups, 2):
        t = stats["学员数"]
        h_rate = f"{stats['高意向']/t*100:.1f}%" if t else "0%"
        r_rate = f"{stats['风险']/t*100:.1f}%" if t else "0%"
        values = [gp, t, stats["高意向"], h_rate, stats["中意向"], stats["观望"], stats["风险"], r_rate]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = data_align
            cell.border = thin_border
            cell.font = Font(name="微软雅黑", size=10)

    # Auto-filter
    ws.auto_filter.ref = f"A1:H{len(sorted_groups) + 1}"
    ws.freeze_panes = "A2"

    wb.save(str(out_excel))
    print(f"[+] 小组统计 Excel: {out_excel}")

    # 5. 输出报告
    out_report = out_dir / f"group_report_{today}.md"
    lines = []
    lines.append(f"# 续费意向 — 小组维度分析")
    lines.append(f"")
    lines.append(f"> 生成时间: {today}")
    lines.append(f"> 总学员数: {len(df_intent)}")
    lines.append(f"")

    # 总览
    total = len(df_intent)
    high_total = sum(s["高意向"] for s in group_stats.values())
    risk_total = sum(s["风险"] for s in group_stats.values())
    lines.append(f"## 一、总览")
    lines.append(f"")
    lines.append(f"| 指标 | 数值 |")
    lines.append(f"|------|------|")
    lines.append(f"| 总学员 | {total} |")
    lines.append(f"| 小组数 | {len(group_stats)} |")
    lines.append(f"| 高意向 | {high_total} ({high_total/total*100:.1f}%) |")
    lines.append(f"| 风险 | {risk_total} ({risk_total/total*100:.1f}%) |")
    lines.append(f"")

    # 小组明细
    lines.append(f"## 二、小组明细（按学员数排序）")
    lines.append(f"")
    lines.append(f"| 小组 | 学员数 | 高意向 | 高意向率 | 中意向 | 观望 | 风险 | 风险率 |")
    lines.append(f"|------|--------|--------|---------|--------|------|------|--------|")
    for gp, stats in sorted_groups:
        t = stats["学员数"]
        h_rate = f"{stats['高意向']/t*100:.1f}%" if t else "0%"
        r_rate = f"{stats['风险']/t*100:.1f}%" if t else "0%"
        lines.append(f"| {gp} | {t} | {stats['高意向']} | {h_rate} | {stats['中意向']} | {stats['观望']} | {stats['风险']} | {r_rate} |")

    lines.append(f"")
    lines.append(f"---")
    lines.append(f"> 由 vipthink-renewal-intention group_stats 生成")

    out_report.write_text("\n".join(lines), encoding="utf-8")
    print(f"[+] 小组报告: {out_report}")
    print(f"[+] 全部完成！")


if __name__ == "__main__":
    main()
