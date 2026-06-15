#!/usr/bin/env python3
"""续费过程格式报告 v4: 团队×池子 核心池展开+退费/结课占比

格式（两行表头）:
  行1: 团队 | 池子 | ←── 执行中 ──→ | 退费占比 | 结课占比
  行2:      |       | 学员|签单率|未续|覆盖|高|中|低|风险|>7天 |         |
  
左侧9列: 核心池(=执行中+停课+等班)展开子指标
右侧2列: 非核心状态占比(基于总人数)
"""
import argparse, os, sys
from collections import defaultdict
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

POOLS = ["升舱", "早鸟", "当月结课", "次月结课", "次次月结课", "活跃低课时"]
POOL_MAP = {
    "升舱未续": "升舱", "早鸟": "早鸟",
    "当月结课未续": "当月结课", "次月结课未续": "次月结课",
    "次次月结课未续": "次次月结课", "活跃低课时未续": "活跃低课时",
}
CORE_STATUS = ["执行中", "停课", "等班"]
TARGET_GROUPS = ["港澳1组", "港澳2组", "港澳组", "美澳1组", "美澳2组", "美澳3组", "美澳4组", "美澳5组"]

# 14 columns: A=团队 B=池子 C=学员 D=签单率 E=未续 F=覆盖 G=高意向 H=中意向 I=低 J=风险 K=>7天 L=断联率 M=退费占比 N=结课占比
G1_HEADERS = ["团队", "池子", "学员", "签单率", "未续", "覆盖率",
              "高意向占比", "中意向占比", "低意向占比", "风险占比",
              "超7天失联率", "断联率", "退费占比", "结课占比"]
COL_WIDTHS = [13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 11.5, 10, 13, 13]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("raw_excel")
    parser.add_argument("--output-dir", default=".")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.raw_excel, data_only=True)
    ws = wb.active

    header_row = 6
    for r in range(1, min(12, ws.max_row + 1)):
        rv = [str(c.value).strip() if c.value else "" for c in ws[r]]
        if "学员ID" in rv and "学员状态" in rv:
            header_row = r; break

    hdrs = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
    sid_c = next((i for i,h in enumerate(hdrs) if h=="学员ID"), 0)
    group_c = next((i for i,h in enumerate(hdrs) if h=="小组"), 3)
    pool_c = next((i for i,h in enumerate(hdrs) if h=="池子"), 15)
    status_c = next((i for i,h in enumerate(hdrs) if h=="学员状态"), 13)
    signed_c = next((i for i,h in enumerate(hdrs) if "是否续费" in h), 8)
    contact_c = next((i for i,h in enumerate(hdrs) if h=="沟通时间"), 18)
    exec_c = next((i for i,h in enumerate(hdrs) if h=="是否执行"), 6)

    today = date.today()
    students = {}
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        sid = str(row[sid_c]).strip().rstrip(".0") if row[sid_c] else ""
        if not sid or sid == "None": continue
        if sid not in students:
            raw_pool = str(row[pool_c]).strip() if row[pool_c] else ""
            students[sid] = {
                "sid": sid,
                "group": str(row[group_c]).strip() if row[group_c] else "",
                "pool": POOL_MAP.get(raw_pool, raw_pool),
                "status": str(row[status_c]).strip() if row[status_c] else "执行中",
                "signed": 0, "has_action": False, "latest_contact": None,
            }
        s = students[sid]
        if row[signed_c] == 1.0: s["signed"] = 1
        if row[exec_c] == 1.0: s["has_action"] = True
        ct = row[contact_c]
        if ct:
            try:
                d = datetime.strptime(str(ct)[:10], "%Y-%m-%d").date()
                if s["latest_contact"] is None or d > s["latest_contact"]:
                    s["latest_contact"] = d
            except: pass
    wb.close()

    # Intent data
    intent_map = {}
    intent_path = os.path.join(args.output_dir, f"renewal_intention_{today}.xlsx")
    if os.path.exists(intent_path):
        iwb = openpyxl.load_workbook(intent_path, data_only=True)
        iws = iwb.active
        ih = [str(c.value).strip() if c.value else "" for c in iws[1]]
        isid = next((i for i,h in enumerate(ih) if "学员ID" in h), 1)
        ilv = next((i for i,h in enumerate(ih) if "意向等级" in h), 5)
        for row in iws.iter_rows(min_row=2, values_only=True):
            sid = str(row[isid]).strip().rstrip(".0") if row[isid] else ""
            lv = str(row[ilv]).strip() if row[ilv] else "LOW"
            if sid: intent_map[sid] = lv
        iwb.close()

    # Output Excel
    excel_path = os.path.join(args.output_dir, f"process_report_{today}.xlsx")
    owb = openpyxl.Workbook()
    ows = owb.active
    ows.title = "过程分析表"

    hfill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    hfont = Font(bold=True, size=11)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    data_align = Alignment(horizontal="center", vertical="center")

    # Row 1: group headers  (A=团队 B=池子 C-L=执行中 M=退费占比 N=结课占比)
    for ci, val in enumerate([("团队", 1), ("池子", 2), ("执行中", 3), ("", 4), ("", 5), ("", 6),
                               ("", 7), ("", 8), ("", 9), ("", 10), ("", 11), ("", 12),
                               ("退费占比", 13), ("结课占比", 14)], 1):
        if val[0]:
            c = ows.cell(row=1, column=ci, value=val[0])
            c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, thin_border
        else:
            c = ows.cell(row=1, column=ci)
            c.fill, c.border = hfill, thin_border

    # Row 2: sub-headers
    for ci, h in enumerate(G1_HEADERS, 1):
        c = ows.cell(row=2, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, thin_border

    # Merge header cells
    ows.merge_cells("A1:A2")   # 团队
    ows.merge_cells("B1:B2")   # 池子
    ows.merge_cells("C1:L1")   # 执行中 (spans C-L, incl 断联率)
    ows.merge_cells("M1:M2")   # 退费占比
    ows.merge_cells("N1:N2")   # 结课占比

    for ci, w in enumerate(COL_WIDTHS, 1):
        ows.column_dimensions[get_column_letter(ci)].width = w

    # Colors
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    table_rows = []
    row_data = []  # (team, pool_rows) pairs for deferred merge
    ri = 3

    for g in TARGET_GROUPS:
        g_students = [s for s in students.values() if s["group"] == g]
        if not g_students: continue

        pool_rows = []
        for p in ["合计"] + POOLS:
            ss = g_students if p == "合计" else [s for s in g_students if s["pool"] == p]
            if not ss: continue

            total = len(ss)
            core = [s for s in ss if s["status"] in CORE_STATUS]
            core_n = len(core)

            refund_n = sum(1 for s in ss if s["status"] == "退费")
            finish_n = sum(1 for s in ss if s["status"] == "结课")
            refund_pct = f"{refund_n/max(total,1)*100:.1f}%"
            finish_pct = f"{finish_n/max(total,1)*100:.1f}%"

            if core_n == 0:
                row = {"团队": g, "池子": p,
                       "学员": 0, "签单率": "-", "未续": 0, "覆盖率": "-",
                       "高": "-", "中": "-", "低": "-", "风险": "-", ">7天": "-", "断联率": "-",
                       "退费": refund_pct, "结课": finish_pct}
                pool_rows.append(row)
                continue

            signed = sum(1 for s in core if s["signed"])
            covered = sum(1 for s in core if s["has_action"])
            lost = sum(1 for s in core if s["latest_contact"] and (today-s["latest_contact"]).days > 7)
            disconnected = sum(1 for s in core if s["has_action"] and s["latest_contact"] and (today-s["latest_contact"]).days > 7)
            disconnect_rate = f"{disconnected/max(covered,1)*100:.1f}%" if covered > 0 else "-"
            high = med = low = risk = 0
            for s in core:
                lv = intent_map.get(s["sid"], "LOW")
                if "HIGH" in lv or "高" in lv: high += 1
                elif "MED" in lv or "中" in lv: med += 1
                elif "RISK" in lv or "风险" in lv: risk += 1
                else: low += 1

            row = {
                "团队": g, "池子": p,
                "学员": core_n,
                "签单率": f"{signed/max(core_n,1)*100:.1f}%",
                "未续": core_n - signed,
                "覆盖率": f"{covered/max(core_n,1)*100:.1f}%",
                "高": f"{high/max(core_n,1)*100:.1f}%",
                "中": f"{med/max(core_n,1)*100:.1f}%",
                "低": f"{low/max(core_n,1)*100:.1f}%",
                "风险": f"{risk/max(core_n,1)*100:.1f}%",
                ">7天": f"{lost/max(core_n,1)*100:.1f}%",
                "断联率": disconnect_rate,
                "退费": refund_pct, "结课": finish_pct,
            }
            pool_rows.append(row)

        row_data.append((ri, pool_rows))
        for pi, row in enumerate(pool_rows):
            vals = [
                row["团队"] if pi == 0 else "",
                row["池子"],
                row["学员"], row["签单率"], row["未续"], row["覆盖率"],
                row["高"], row["中"], row["低"], row["风险"], row[">7天"], row["断联率"],
                row["退费"], row["结课"],
            ]
            for ci, v in enumerate(vals, 1):
                c = ows.cell(row=ri, column=ci, value=v)
                c.border = thin_border
                c.alignment = data_align

            try:
                v7 = float(row[">7天"].replace("%",""))
                if v7 >= 30:
                    ows.cell(row=ri, column=11).fill = red_fill
                elif v7 >= 20:
                    ows.cell(row=ri, column=11).fill = yellow_fill
            except: pass
            try:
                cv = float(row["覆盖率"].replace("%",""))
                if cv < 25:
                    ows.cell(row=ri, column=6).fill = red_fill
                elif cv < 40:
                    ows.cell(row=ri, column=6).fill = yellow_fill
            except: pass

            table_rows.append(row)
            ri += 1

        ri += 1  # group separator

    # Merge team name cells AFTER writing all data
    for start_r, pool_rows in row_data:
        n = len(pool_rows)
        if n > 1:
            ows.merge_cells(start_row=start_r, start_column=1, end_row=start_r+n-1, end_column=1)

    ows.freeze_panes = "A3"
    owb.save(excel_path)

    if args.json:
        import json
        print(json.dumps({"status": "ok", "path": excel_path, "rows": table_rows}, ensure_ascii=False))
    else:
        print(f"过程报告: {excel_path}")
        for r in table_rows:
            print(f"  {r['团队']:<8s} {r['池子']:<8s} | "
                  f"学员{r['学员']:>4d} | {r['签单率']:>6s} | 未续{r['未续']:>4d} | "
                  f"覆盖{r['覆盖率']:>5s} | H{r['高']:>5s} M{r['中']:>5s} "
                  f"L{r['低']:>5s} R{r['风险']:>5s} | >7天{r['>7天']:>5s} | "
                  f"退费{r['退费']:>5s} 结课{r['结课']:>5s}")

    return table_rows


if __name__ == "__main__":
    main()
